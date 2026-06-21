"""006 表格处理工具 — 阶段0 Demo
多源异构社保数据 → 按人(姓名+证件)横向归集到标准字段模板。
只读源文件,输出 _合并结果.xlsx + _异常清单.xlsx,不动原件。

验证目标(spec §8):解析正确 / 映射正确 / 按人合并正确 / 完整性对账 / 异常可见 / 自动识别准确率。
"""
from __future__ import annotations
import glob, os, sys, json, warnings
from collections import defaultdict
import openpyxl

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8")

SRC_DIR = r"D:\乐逗\Desktop\社保处理文件"

# ── 标准字段(目标模板,纯配置) ───────────────────────────────
STD_FIELDS = [
    "养老个人", "养老公司", "医疗个人", "医疗公司",
    "失业个人", "失业公司", "工伤公司", "生育公司",
    "大额互助个人", "大额互助公司",
    "公积金个人", "公积金公司", "服务费", "残保金",
]
MERGE_KEYS = ["姓名", "证件号码"]  # 归集主键(可配)

# ── 源映射配置(每种结构一份) ────────────────────────────────
# match: 表头特征(子集命中即认);sheet: 目标sheet关键词/None=第一个;header: 表头行区间(1-based)
# key_map: 主键列(源→标准);col_map: 源有效表头列名 → 标准字段
MAPPINGS = [
    {
        "id": "sz_daily_declare", "name": "深圳-日常申报明细",
        "match": ["序号", "姓名", "证件号码", "费款所属期起", "缴费工资"],
        "sheet_kw": "申报明细", "header": (1, 2),
        "key_map": {"姓名": "姓名", "证件号码": "证件号码"},
        "col_map": {
            "基本养老保险（个人）/应缴费额": "养老个人", "基本养老保险（单位）/应缴费额": "养老公司",
            "基本医疗保险（个人）/应缴费额": "医疗个人", "基本医疗保险（单位）/应缴费额": "医疗公司",
            "失业保险（个人）/应缴费额": "失业个人", "失业保险（单位）/应缴费额": "失业公司",
            "工伤保险（单位）/应缴费额": "工伤公司", "生育保险/应缴费额": "生育公司",
        },
    },
    {
        "id": "qqxb_shebao", "name": "亲亲小保托管-社保sheet",
        "match": ["序号", "姓名", "证件号码", "缴费工资", "公司"],
        "sheet_kw": "社保", "header": (1, 2),
        "key_map": {"姓名": "姓名", "证件号码": "证件号码"},
        "col_map": {
            "基本养老保险（个人）/应缴费额": "养老个人", "基本养老保险（单位）/应缴费额": "养老公司",
            "基本医疗保险（个人）/应缴费额": "医疗个人", "基本医疗保险（单位）/应缴费额": "医疗公司",
            "失业保险（个人）/应缴费额": "失业个人", "失业保险（单位）/应缴费额": "失业公司",
            "工伤保险（单位）/应缴费额": "工伤公司", "生育保险/应缴费额": "生育公司",
        },
    },
    {
        "id": "sz_housing_single", "name": "深圳-单笔缴存清单(公积金)",
        "match": ["姓名", "证件号码", "个人账号", "缴存基数（元）", "金额合计（元）"],
        "sheet_kw": "单笔缴存", "header": (4, 4),
        "key_map": {"姓名": "姓名", "证件号码": "证件号码"},
        "col_map": {},
        "derived_fields": [
            {"target": "公积金个人", "expr": "{缴存基数（元）}*{个人缴存比例}", "round": 2},
            {"target": "公积金公司", "expr": "{缴存基数（元）}*{单位缴存比例}", "round": 2},
        ],
        "derive_check": {"sum_of": ["公积金个人", "公积金公司"], "equals_col": "金额合计（元）", "tol": 0.05},
    },
    {
        "id": "qqxb_housing", "name": "亲亲小保托管-公积金sheet",
        "match": ["姓名", "证件号码", "个人账号", "缴存基数（元）", "公司"],
        "sheet_kw": "公积金", "header": (2, 2),
        "key_map": {"姓名": "姓名", "证件号码": "证件号码"},
        "col_map": {},
        "derived_fields": [
            {"target": "公积金个人", "expr": "{缴存基数（元）}*{个人缴存比例}", "round": 2},
            {"target": "公积金公司", "expr": "{缴存基数（元）}*{单位缴存比例}", "round": 2},
        ],
        "derive_check": {"sum_of": ["公积金个人", "公积金公司"], "equals_col": "金额合计（元）", "tol": 0.05},
    },
    {
        "id": "hk_mpf", "name": "香港强积金",
        "match": ["Name", "Member Account No", "ID No.", "Total Contribution Amount"],
        "sheet_kw": None, "header": (1, 1),
        "key_map": {"Name": "姓名", "ID No.": "证件号码"},
        "col_map": {
            "Employee mandatory contributions": "公积金个人",
            "Employer mandatory contributions": "公积金公司",
        },
    },
    {
        "id": "yidi_shebao", "name": "异地代缴-社保明细",
        "match": ["员工", "身份证号码", "实际缴费单位", "社保应缴金额"],
        "sheet_kw": "社保代缴明细表", "header": (1, 2),
        "key_map": {"员工": "姓名", "身份证号码": "证件号码"},
        "col_map": {
            "养老/养老个人": "养老个人", "养老/养老单位": "养老公司",
            "医疗/医疗个人": "医疗个人", "医疗/医疗单位": "医疗公司",
            "失业/失业个人": "失业个人", "失业/失业单位": "失业公司",
            "工伤/工伤单位": "工伤公司", "生育/生育单位": "生育公司",
            "残保金/残保金合计金额": "残保金",
        },
    },
    {
        "id": "yidi_housing", "name": "异地代缴-公积金明细",
        "match": ["员工", "身份证号码", "公积金应缴金额", "公积金/公积金个人"],
        "sheet_kw": "公积金代缴明细表", "header": (1, 2),
        "key_map": {"员工": "姓名", "身份证号码": "证件号码"},
        "col_map": {"公积金/公积金个人": "公积金个人", "公积金/公积金单位": "公积金公司"},
    },
    {
        "id": "yidi_service", "name": "异地代缴-服务费",
        "match": ["姓名", "身份证号码", "服务月份", "金额"],
        "sheet_kw": "人事代理服务费", "header": (1, 1),
        "key_map": {"姓名": "姓名", "身份证号码": "证件号码"},
        "col_map": {"金额": "服务费"},
    },
    {
        "id": "bj_shebao", "name": "北京分公司-社保",
        "match": ["序号", "姓名", "身份证号", "应收金额/应收合计"],
        "sheet_kw": "社保缴交明细", "header": (2, 3),
        "key_map": {"姓名": "姓名", "身份证号": "证件号码"},
        "col_map": {
            "养老保险/个人交": "养老个人", "养老保险/单位交": "养老公司",
            "医疗及生育保险/个人交": "医疗个人", "医疗及生育保险/单位交": "医疗公司",
            "失业保险/个人交": "失业个人", "失业保险/单位交": "失业公司",
            "工伤保险/单位交": "工伤公司",
            "大额互助资金/个人交": "大额互助个人", "大额互助资金/单位交": "大额互助公司",
        },
    },
    {
        "id": "bj_housing", "name": "北京分公司-公积金",
        "match": ["姓名", "证件号码", "个人缴存基数", "月缴存额合计"],
        "sheet_kw": "公积金", "header": (5, 5),
        "key_map": {"姓名": "姓名", "证件号码": "证件号码"},
        "col_map": {"个人月缴存额": "公积金个人", "单位月缴存额": "公积金公司"},
    },
]


def fill_merged(ws):
    vals = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                vals[(r, c)] = v
    return vals


def parse_header(ws, r0, r1):
    m = fill_merged(ws)
    cols = []
    for c in range(1, ws.max_column + 1):
        parts = []
        for r in range(r0, r1 + 1):
            v = m.get((r, c), ws.cell(r, c).value)
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
        cols.append("/".join(dict.fromkeys(parts)) if parts else None)
    return cols


def pick_sheet(wb, kw):
    if not kw:
        return wb.worksheets[0]
    for ws in wb.worksheets:
        if kw in ws.title:
            return ws
    return wb.worksheets[0]


def match_mapping(headers_by_sheet):
    """返回 [(mapping, sheet_title)] — 一个文件可命中多个(多sheet)。"""
    hits = []
    for mp in MAPPINGS:
        for title, hdr in headers_by_sheet.items():
            if mp["sheet_kw"] and mp["sheet_kw"] not in title:
                continue
            hset = set(h for h in hdr if h)
            need = mp["match"]
            ok = sum(1 for k in need if k in hset)
            if ok >= max(3, len(need) - 1):  # 子集命中(允许缺1)
                hits.append((mp, title, ok / len(need)))
    # 同一 sheet 只保留最佳命中
    best = {}
    for mp, title, score in hits:
        if title not in best or score > best[title][2]:
            best[title] = (mp, title, score)
    return list(best.values())


def eval_expr(expr, getval):
    """通用表达式求值:{源列名} 占位 → 数值,仅支持 + - * / 与括号。"""
    import re as _re
    def repl(m):
        v = getval(m.group(1))
        try:
            return repr(float(v))
        except (ValueError, TypeError):
            raise ValueError(f"列[{m.group(1)}]非数值: {v!r}")
    safe = _re.sub(r"\{([^{}]+)\}", repl, expr)
    if not _re.fullmatch(r"[0-9eE+\-*/.() ]+", safe):
        raise ValueError(f"表达式含非法字符: {safe}")
    return eval(safe, {"__builtins__": {}}, {})


def is_skip_row(rowvals, key_idx):
    """跳过合计/空行:主键列为空,或含合计字样。"""
    kv = [rowvals[i] for i in key_idx if i is not None and i < len(rowvals)]
    if not any(v is not None and str(v).strip() for v in kv):
        return True
    for v in kv:
        if v and any(t in str(v) for t in ("合计", "小计", "总计")):
            return True
    return False


def main():
    files = sorted(glob.glob(os.path.join(SRC_DIR, "*.xlsx")))
    files = [f for f in files if not os.path.basename(f).startswith(("_", "~$"))]
    person = {}            # (姓名,证件) -> {标准字段:值}
    person_src = defaultdict(list)
    anomalies = []         # 异常清单
    recognize_log = []     # 自动识别日志
    src_record_count = 0
    field_sum_check = defaultdict(float)  # 各标准字段来源累加(对账)

    for f in files:
        fname = os.path.basename(f)
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception as e:
            anomalies.append(["读取失败", fname, "", str(e)]); continue
        headers_by_sheet = {}
        # 先用每个 sheet 各自配置里的表头行试解析(取并集特征)
        for ws in wb.worksheets:
            if ws.max_row <= 1:
                continue
            # 用候选 header 行解析:尝试所有 mapping 的 header 区间收集特征
            for r0r1 in {mp["header"] for mp in MAPPINGS}:
                hdr = parse_header(ws, *r0r1)
                headers_by_sheet[(ws.title, r0r1)] = hdr
        # 匹配:按 (sheet,header区间) 找命中
        matched = []
        for (title, r0r1), hdr in headers_by_sheet.items():
            for mp in MAPPINGS:
                if mp["sheet_kw"] and mp["sheet_kw"] not in title:
                    continue
                if mp["header"] != r0r1:
                    continue
                hset = set(h for h in hdr if h)
                ok = sum(1 for k in mp["match"] if k in hset)
                if ok >= max(3, len(mp["match"]) - 1):
                    matched.append((mp, title, hdr, ok / len(mp["match"])))
        # 去重:同 sheet 取最佳
        best = {}
        for mp, title, hdr, score in matched:
            if title not in best or score > best[title][3]:
                best[title] = (mp, title, hdr, score)
        if not best:
            anomalies.append(["未命中任何模板", fname, "", "需AI生成映射草稿(②)"])
            recognize_log.append([fname, "—", "无命中", ""])
            wb.close(); continue

        for title, (mp, _t, hdr, score) in best.items():
            recognize_log.append([fname, title, mp["name"], f"{score:.0%}"])
            ws = wb[title]
            col_idx = {h: i for i, h in enumerate(hdr) if h}
            key_idx = [col_idx.get(k) for k in mp["key_map"]]
            data_start = mp["header"][1] + 1
            for r in range(data_start, ws.max_row + 1):
                rowvals = [ws.cell(r, c + 1).value for c in range(len(hdr))]
                if is_skip_row(rowvals, [i for i in key_idx if i is not None]):
                    continue
                # 主键
                key = {}
                for srcname, stdname in mp["key_map"].items():
                    i = col_idx.get(srcname)
                    key[stdname] = str(rowvals[i]).strip() if i is not None and rowvals[i] is not None else ""
                pk = tuple(key.get(k, "") for k in MERGE_KEYS)
                if not pk[0] and not pk[1]:
                    continue
                src_record_count += 1
                rec = person.setdefault(pk, {k: None for k in MERGE_KEYS})
                rec["姓名"], rec["证件号码"] = pk
                person_src[pk].append(f"{mp['name']}")

                def add_field(stdname, val):
                    if isinstance(val, (int, float)):
                        prev = rec.get(stdname)
                        rec[stdname] = (prev or 0) + val if isinstance(prev, (int, float)) else val
                        field_sum_check[stdname] += val
                    else:
                        rec[stdname] = val

                # 业务字段(直接映射)
                for srcname, stdname in mp["col_map"].items():
                    i = col_idx.get(srcname)
                    if i is None:
                        continue
                    val = rowvals[i]
                    if val is None or str(val).strip() == "":
                        continue
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        pass
                    add_field(stdname, val)

                # 派生字段(通用表达式,按配置拆列)
                def getcol(name):
                    i = col_idx.get(name)
                    return rowvals[i] if i is not None else None
                derived = {}
                for d in mp.get("derived_fields", []):
                    try:
                        v = round(eval_expr(d["expr"], getcol), d.get("round", 2))
                    except ValueError as e:
                        anomalies.append(["派生失败", fname, f"{pk} {d['target']}", str(e)]); continue
                    derived[d["target"]] = v
                    add_field(d["target"], v)
                # 派生校验:拆分项之和 == 原合计列(容差内)
                chk = mp.get("derive_check")
                if chk and all(t in derived for t in chk["sum_of"]):
                    i = col_idx.get(chk["equals_col"])
                    total = getcol(chk["equals_col"])
                    try:
                        total = float(total)
                        s = sum(derived[t] for t in chk["sum_of"])
                        if abs(s - total) > chk.get("tol", 0.05):
                            anomalies.append(["拆分校验不符", fname, f"{pk}", f"拆分和{s} vs 合计{total}"])
                    except (ValueError, TypeError):
                        pass
        wb.close()

    # ── 输出合并结果 ──────────────────────────────
    out_cols = MERGE_KEYS + STD_FIELDS + ["来源"]
    wb_out = openpyxl.Workbook()
    ws = wb_out.active; ws.title = "归集结果"
    ws.append(out_cols)
    for pk, rec in sorted(person.items()):
        row = [rec.get(c) for c in MERGE_KEYS + STD_FIELDS]
        row.append(" + ".join(sorted(set(person_src[pk]))))
        ws.append(row)
    out_path = os.path.join(SRC_DIR, "_合并结果.xlsx")
    wb_out.save(out_path)

    # ── 输出异常+识别清单 ──────────────────────────
    wb_a = openpyxl.Workbook()
    wsa = wb_a.active; wsa.title = "识别日志"
    wsa.append(["文件", "sheet", "命中模板", "特征匹配度"])
    for r in recognize_log: wsa.append(r)
    wsx = wb_a.create_sheet("异常清单")
    wsx.append(["类型", "文件", "定位", "详情"])
    for r in anomalies: wsx.append(r)
    a_path = os.path.join(SRC_DIR, "_识别与异常.xlsx")
    wb_a.save(a_path)

    # ── 控制台对账摘要 ─────────────────────────────
    print("="*60)
    print(f"源文件数: {len(files)}  |  解析记录数: {src_record_count}  |  归集后人数: {len(person)}")
    print(f"异常数: {len(anomalies)}")
    print("\n[自动识别结果]")
    for r in recognize_log:
        print(f"  {r[3]:>5}  {r[2]:<22} ← {r[0][:30]} [{r[1]}]")
    print("\n[各标准字段累计金额(对账用)]")
    for k in STD_FIELDS:
        if field_sum_check.get(k):
            print(f"  {k}: {field_sum_check[k]:,.2f}")
    if anomalies:
        print("\n[异常样本(前10)]")
        for r in anomalies[:10]:
            print("  ", r)
    print(f"\n输出: {out_path}")
    print(f"输出: {a_path}")


if __name__ == "__main__":
    main()
