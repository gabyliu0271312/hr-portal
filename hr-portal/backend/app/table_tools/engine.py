"""通用归集引擎 —— 零业务字段。

只认抽象动作:解析多行表头(合并填充)、按映射搬列、按表达式派生、
按主键归集、按口径聚合、校验、标来源。完全不知道"养老""公积金"为何物。

所有业务语义来自运行时传入的 mapping 配置(模板库),不写死在此。

派生字段求值复用公共公式引擎 app.ai_formula(报表/数据集同款),
支持 IF/ROUND/MIN/MAX/SUM/CALC_TAX 等全套函数。派生表达式用 {列名} 占位,
内部转成引擎原生的 FIELD("列名") 求值。
"""
from __future__ import annotations

import re
from collections import defaultdict
from typing import Any, Callable

import openpyxl

from app.ai_formula.formula_evaluator import evaluate_formula


# ── 表头解析 ────────────────────────────────────────────────
def _fill_merged(ws) -> dict[tuple[int, int], Any]:
    """合并单元格:把左上角值填充到区域内每个坐标。"""
    vals: dict[tuple[int, int], Any] = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                vals[(r, c)] = v
    return vals


def parse_header(ws, start_row: int, end_row: int) -> list[str | None]:
    """解析 [start_row, end_row] 区间的多行表头,逐列拼接去重。

    返回每列的有效表头名(空列为 None)。1-based 行号。
    """
    merged = _fill_merged(ws)
    cols: list[str | None] = []
    for c in range(1, ws.max_column + 1):
        parts: list[str] = []
        for r in range(start_row, end_row + 1):
            v = merged.get((r, c), ws.cell(r, c).value)
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
        cols.append("/".join(dict.fromkeys(parts)) if parts else None)
    return cols


def sheet_headers(ws, header_rows_candidates: set[tuple[int, int]]) -> dict[tuple[int, int], list[str | None]]:
    """对一个 sheet,按多组候选表头行区间各解析一次(供识别匹配用)。"""
    return {hr: parse_header(ws, *hr) for hr in header_rows_candidates}


# ── 通用表达式求值(派生字段) ───────────────────────────────
_PLACEHOLDER_RE = re.compile(r"\{([^{}]+)\}")


def _to_field_calls(expr: str) -> tuple[str, list[str]]:
    """把派生表达式里的 {列名} 占位转成引擎原生 FIELD("列名")。

    返回 (转换后表达式, 引用的列名列表)。列名原样保留(含中文/括号/斜杠),
    FIELD 的参数是字面量字符串,引擎按字符串向 resolver 取值,无需归一化。
    """
    refs: list[str] = []

    def repl(m: re.Match) -> str:
        name = m.group(1).strip()
        refs.append(name)
        # 列名可能含双引号(罕见),用单引号包裹规避;FIELD 正则两种引号都认
        quote = "'" if '"' in name else '"'
        return f"FIELD({quote}{name}{quote})"

    return _PLACEHOLDER_RE.sub(repl, expr), refs


def eval_derived(
    expr: str,
    getval: Callable[[str], Any],
    custom_functions: dict[str, Callable[..., Any]] | None = None,
    *,
    missing_as_zero: bool = False,
) -> Any:
    """派生字段求值:{列名} 占位 → 公共公式引擎。"""
    converted, refs = _to_field_calls(expr)
    if not missing_as_zero and any(getval(name) in (None, "") for name in refs):
        return None
    resolver = getval
    if missing_as_zero:
        resolver = lambda name: getval(name) if getval(name) not in (None, "") else 0
    return evaluate_formula(
        converted,
        field_resolver=resolver,
        custom_functions=custom_functions,
    )


# ── 行级工具 ────────────────────────────────────────────────
def is_skip_row(rowvals: list[Any], key_idx: list[int], skip_tokens: list[str]) -> bool:
    """跳过合计/空行:主键列全空,或含合计字样。"""
    kv = [rowvals[i] for i in key_idx if i is not None and i < len(rowvals)]
    if not any(v is not None and str(v).strip() for v in kv):
        return True
    for v in kv:
        if v and any(t in str(v) for t in skip_tokens):
            return True
    return False


# ── 单 sheet 解析为标准记录 ─────────────────────────────────
def extract_records(
    ws,
    header: list[str | None],
    mapping: dict,
    custom_functions: dict[str, Callable[..., Any]] | None = None,
) -> tuple[list[dict], list[dict]]:
    """按一份 source_mapping 把一个 sheet 解析成标准记录列表。

    mapping 关键字段:
      key_map      源列→标准主键列
      column_map   源列→标准字段(直接搬)
      derived_fields  [{target, expr, round}]  派生(公共公式引擎,{列名} 占位)
      derive_check {sum_of, equals_col, tol}   拆分校验
      header (start,end)  表头行区间
      skip_tokens  合计行关键词
    custom_functions: 公共引擎可执行函数库(IF/ROUND/CALC_TAX 等),由调用方注入。
    返回 (records, anomalies)
    """
    col_idx = {h: i for i, h in enumerate(header) if h}
    key_map: dict[str, str] = mapping["key_map"]
    skip_tokens = mapping.get("skip_tokens", ["合计", "小计", "总计"])
    key_idx = [col_idx.get(k) for k in key_map]
    data_start = mapping["header"][1] + 1

    records: list[dict] = []
    anomalies: list[dict] = []

    def getcol(name: str):
        i = col_idx.get(name)
        return ws.cell(_r, i + 1).value if i is not None else None

    for _r in range(data_start, ws.max_row + 1):
        rowvals = [ws.cell(_r, c + 1).value for c in range(len(header))]
        if is_skip_row(rowvals, [i for i in key_idx if i is not None], skip_tokens):
            continue
        rec: dict[str, Any] = {}
        # 主键
        for srcname, stdname in key_map.items():
            i = col_idx.get(srcname)
            if i is None or rowvals[i] is None:
                rec[stdname] = ""
            else:
                v = rowvals[i]
                # Excel 常把身份证号存成数字(float)，str() 会得到科学计数法，需转回整数字符串
                if isinstance(v, float) and v == int(v):
                    rec[stdname] = str(int(v))
                else:
                    rec[stdname] = str(v).strip()
        if not any(rec.get(v) for v in key_map.values()):
            continue
        # 直接映射（跳过已被 key_map 赋值的主键列，避免 float() 覆盖字符串主键）
        key_std_fields = set(key_map.values())
        for srcname, stdname in mapping.get("column_map", {}).items():
            if stdname in key_std_fields:
                continue
            i = col_idx.get(srcname)
            if i is None:
                continue
            val = rowvals[i]
            if val is None or str(val).strip() == "":
                continue
            try:
                val = float(val)
            except (ValueError, TypeError):
                pass
            rec[stdname] = val
        # 派生字段统一在按主键聚合后,使用标准字段计算
        # 拆分校验也在聚合后执行
        records.append(rec)
    return records, anomalies


def _normalize_key_value(value: Any) -> str:
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value if value is not None else "").strip()


def _mapping_key(record: dict, fields: list[str]) -> tuple[str, ...]:
    return tuple(_normalize_key_value(record.get(field, "")) for field in fields)


def apply_key_mappings(
    records_with_src: list[tuple[dict, str]],
    merge_keys: list[str],
    key_mappings: list[dict] | None = None,
) -> tuple[list[tuple[dict, str]], dict, list[dict], list[dict]]:
    """应用模板级精确主键映射；未命中时保持原始主键。"""
    entries: dict[tuple[str, ...], dict] = {}
    for item in key_mappings or []:
        source_key = item.get("source_key") or {}
        canonical = item.get("canonical_merge_key") or {}
        if set(source_key) != set(merge_keys) or set(canonical) != set(merge_keys):
            continue
        key = tuple(_normalize_key_value(source_key[field]) for field in merge_keys)
        if key in entries and entries[key] != item:
            raise ValueError("主键映射冲突")
        entries[key] = item

    stats = {"configured": len(entries), "matched": 0, "unmatched": 0}
    anomalies: list[dict] = []
    traces: list[dict] = []
    mapped: list[tuple[dict, str]] = []
    for rec, src in records_with_src:
        source_key = {field: rec.get(field, "") for field in merge_keys}
        item = entries.get(_mapping_key(rec, merge_keys))
        if item is None:
            if entries:
                stats["unmatched"] += 1
            mapped.append((rec, src))
            continue
        stats["matched"] += 1
        canonical = item["canonical_merge_key"]
        for field in merge_keys:
            rec[field] = _normalize_key_value(canonical[field])
        traces.append({"source_key": source_key, "canonical_merge_key": canonical, "source": src})
        mapped.append((rec, src))
    return mapped, stats, anomalies, traces


# ── 按主键归集 + 聚合 ───────────────────────────────────────
def aggregate_records(
    records_with_src: list[tuple[dict, str]],
    merge_keys: list[str],
    std_fields: list[str],
    agg: str = "sum",
    derived_fields: list[dict] | None = None,
    custom_functions: dict[str, Callable[..., Any]] | None = None,
    derived_available_fields: set[str] | None = None,
    fill_missing_fields: bool = True,
) -> tuple[list[dict], list[dict]]:
    """按主键归集标准字段,再基于归集结果计算派生字段。"""
    person: dict[tuple, dict] = {}
    person_src: dict[tuple, list[str]] = defaultdict(list)
    anomalies: list[dict] = []

    for rec, src in records_with_src:
        # 主键统一转字符串:不同来源对同一字段可能写入 str/float 混合类型
        # (如证件号被误归到 column_map 会被 float() 转换),不统一会导致
        # sorted() 报 TypeError,也会导致同一人在不同来源里因类型不同而对不上。
        pk = tuple(str(int(rec.get(k, 0))) if isinstance(rec.get(k), float) and rec.get(k) == int(rec.get(k)) else str(rec.get(k, "")).strip() for k in merge_keys)
        if not any(pk):
            continue
        cur = person.setdefault(pk, {k: rec.get(k, "") for k in merge_keys})
        person_src[pk].append(src)
        for f in std_fields:
            if f not in rec:
                continue
            val = rec[f]
            if val is None or val == "":
                continue
            prev = cur.get(f)
            if isinstance(val, (int, float)):
                if isinstance(prev, (int, float)):
                    if agg == "conflict" and abs(prev - val) > 0.01:
                        anomalies.append({"type": "金额冲突", "key": dict(zip(merge_keys, pk)),
                                          "detail": f"{f}: {prev} vs {val}"})
                    else:
                        cur[f] = prev + val
                else:
                    cur[f] = val
            else:
                if prev in (None, ""):
                    cur[f] = val

    rows: list[dict] = []
    for pk in sorted(person):
        row = dict(person[pk])
        row["来源"] = " + ".join(sorted(set(person_src[pk])))
        # 派生字段:仅当公式引用的标准字段都真实存在时才计算,
        # 否则跳过、保留该字段已有的直接映射值(如来源表直接提供)。
        for d in derived_fields or []:
            _, refs = _to_field_calls(d["expr"])
            available = derived_available_fields if derived_available_fields is not None else set(row)
            if any(ref not in available for ref in refs):
                continue
            # If this row has none of the dependencies, preserve an existing
            # direct-mapped target from another source instead of overwriting it.
            if refs and not any(row.get(ref) not in (None, "") for ref in refs) and d["target"] in row:
                continue
            value = eval_derived(d["expr"], row.get, custom_functions, missing_as_zero=True)
            if value is None or value == "":
                continue
            if isinstance(value, (int, float)) and "round" in d:
                value = round(value, d.get("round", 2))
            row[d["target"]] = value
        if fill_missing_fields:
            # 缺失字段填 0(输出口径),须在派生之后执行以免抹掉"缺失"信号。
            for field in std_fields:
                row.setdefault(field, 0)
        rows.append(row)
    return rows, anomalies


def _aggregate_mapping_results(
    records_with_src: list[tuple[dict, str]],
    merge_keys: list[str],
    std_fields: list[str],
    agg: str,
    mappings: list[dict],
    custom_functions: dict[str, Callable[..., Any]] | None = None,
) -> tuple[list[dict], list[dict]]:
    """先在来源映射内派生，再跨映射汇总，避免公式覆盖其他映射的直接值。"""
    mapping_by_name = {mapping["name"]: mapping for mapping in mappings}
    records_by_mapping: dict[str, list[tuple[dict, str]]] = defaultdict(list)
    for record, source in records_with_src:
        records_by_mapping[source].append((record, source))

    mapping_rows: list[tuple[dict, str]] = []
    anomalies: list[dict] = []
    for source, source_records in records_by_mapping.items():
        mapping = mapping_by_name[source]
        available_fields = set(merge_keys)
        available_fields.update((mapping.get("column_map") or {}).values())
        rows, source_anomalies = aggregate_records(
            source_records,
            merge_keys,
            std_fields,
            agg,
            mapping.get("derived_fields") or [],
            custom_functions,
            available_fields,
            fill_missing_fields=False,
        )
        if mapping.get("derive_check"):
            _apply_derive_checks(
                rows, merge_keys, [mapping["derive_check"]], source_anomalies
            )
        anomalies.extend(source_anomalies)
        mapping_rows.extend((row, source) for row in rows)

    rows, aggregate_anomalies = aggregate_records(
        mapping_rows, merge_keys, std_fields, agg
    )
    anomalies.extend(aggregate_anomalies)
    return rows, anomalies


def _merge_derived_fields(mappings: list[dict]) -> list[dict]:
    """合并各来源的模板级派生配置；同一目标必须使用同一口径。"""
    merged: dict[str, dict] = {}
    for mapping in mappings:
        for field in mapping.get("derived_fields", []):
            target = field["target"]
            current = {key: field.get(key) for key in ("target", "expr", "round")}
            if target in merged and merged[target] != current:
                raise ValueError(f"派生字段 {target} 在多个来源映射中的公式不一致")
            merged[target] = current
    return list(merged.values())


def _apply_derive_checks(rows: list[dict], merge_keys: list[str], checks: list[dict], anomalies: list[dict]) -> None:
    """基于归集后的标准字段执行拆分校验。"""
    for row in rows:
        for check in checks:
            try:
                total = float(row.get(check["equals_col"], 0) or 0)
                subtotal = sum(float(row.get(field, 0) or 0) for field in check["sum_of"])
            except (TypeError, ValueError):
                continue
            if abs(subtotal - total) > check.get("tol", 0.05):
                anomalies.append({
                    "type": "拆分校验不符",
                    "key": {field: row.get(field, "") for field in merge_keys},
                    "detail": f"和{subtotal} vs 合计{total}",
                })


# ── 顶层:跑一个合并任务 ────────────────────────────────────
def run_merge(
    files: list[tuple[str, bytes]],
    template: dict,
    mappings: list[dict],
    custom_functions: dict[str, Callable[..., Any]] | None = None,
    key_mappings: list[dict] | None = None,
) -> dict:
    """执行一次合并。

    files: [(filename, xlsx_bytes)]
    template: {merge_keys, std_fields, aggregate}
    mappings: [source_mapping]  每份含 match/sheet_kw/header/key_map/column_map/derived_fields...
    custom_functions: 公共公式引擎可执行函数库(IF/ROUND/CALC_TAX 等),由 router 注入。
    返回 {rows, columns, recognize_log, anomalies, stats}
    """
    import io

    merge_keys = template["merge_keys"]
    std_fields = template["std_fields"]
    agg = template.get("aggregate", "sum")
    _merge_derived_fields(mappings)
    header_candidates = {tuple(m["header"]) for m in mappings}

    records_with_src: list[tuple[dict, str]] = []
    recognize_log: list[dict] = []
    anomalies: list[dict] = []
    record_count = 0

    for fname, blob in files:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
        except Exception as e:
            anomalies.append({"type": "读取失败", "key": fname, "detail": str(e)})
            continue
        for ws in wb.worksheets:
            if ws.max_row <= 1:
                continue
            headers = sheet_headers(ws, header_candidates)
            # 找命中的 mapping(表头特征 + sheet 关键词 + 表头行区间一致)
            best: tuple[dict, float] | None = None
            best_hdr: list | None = None
            is_ambiguous = False
            for m in mappings:
                if m.get("sheet_kw") and m["sheet_kw"] not in ws.title:
                    continue
                hdr = headers.get(tuple(m["header"]))
                if not hdr:
                    continue
                hset = {h for h in hdr if h}
                need = m["match"]
                hit = sum(1 for k in need if k in hset)
                if hit >= max(3, len(need) - 1):
                    score = hit / len(need)
                    if best is None or score > best[1]:
                        best, best_hdr = (m, score), hdr
                        is_ambiguous = False
                    elif best is not None and score == best[1]:
                        is_ambiguous = True
            if best is None:
                anomalies.append({
                    "type": "未匹配源映射",
                    "key": f"{fname} / {ws.title}",
                    "detail": "未找到符合表头特征和 Sheet 规则的源映射",
                    "file": fname,
                })
                continue
            if is_ambiguous:
                anomalies.append({
                    "type": "映射命中歧义",
                    "key": f"{fname} / {ws.title}",
                    "detail": "存在多个同分源映射，已跳过该 Sheet，请维护更稳定的表头特征",
                    "file": fname,
                })
                continue
            m, score = best
            recognize_log.append({"file": fname, "sheet": ws.title,
                                  "mapping": m["name"], "score": round(score, 3)})
            recs, anos = extract_records(ws, best_hdr, m, custom_functions)
            for a in anos:
                a["file"] = fname
            anomalies.extend(anos)
            record_count += len(recs)
            for rec in recs:
                records_with_src.append((rec, m["name"]))
        wb.close()

    records, key_mapping_stats, mapping_anomalies, raw_key_traces = apply_key_mappings(
        records_with_src, merge_keys, key_mappings
    )
    rows, agg_anomalies = _aggregate_mapping_results(
        records, merge_keys, std_fields, agg, mappings, custom_functions
    )
    anomalies.extend(mapping_anomalies)
    anomalies.extend(agg_anomalies)

    columns = merge_keys + std_fields + ["来源"]
    return {
        "rows": rows,
        "columns": columns,
        "recognize_log": recognize_log,
        "anomalies": anomalies,
        "stats": {"files": len(files), "records": record_count, "persons": len(rows),
                  "anomalies": len(anomalies)},
        "key_mapping_stats": key_mapping_stats,
        "raw_key_traces": raw_key_traces,
    }


def rows_to_xlsx(
    columns: list[str],
    rows: list[dict],
    column_headers: list[str] | None = None,
) -> bytes:
    """归集结果导出为 xlsx bytes。"""
    import io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "归集结果"
    ws.append(column_headers or columns)
    for row in rows:
        ws.append([row.get(c) for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
