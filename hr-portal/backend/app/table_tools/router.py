"""table_tools 路由:归集模板库 + 多表合并执行。

权限(固定动作 V/C/U/D/E,owner 隔离写操作 —— 方案 B):
  V = 查看模板 / 用模板跑合并
  C = 新建模板(记 created_by)
  U = 改模板;仅能改自己建的,超级管理员可改任何模板
  D = 删模板;仅能删自己建的,超级管理员可删任何模板
  E = 导出/下载结果
"""
from __future__ import annotations

import hashlib
import io
import json
import re
import secrets
from datetime import datetime, timedelta, timezone
from typing import Any

import openpyxl

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import Response
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.ai.audit import AiAuditTimer, record_ai_log
from app.ai.capabilities import get_capability
from app.ai.policy_guard import enforce_output_deny_patterns, validate_capability_policy
from app.ai_formula.custom_functions import executable_functions
from app.ai_formula.formula_evaluator import formula_syntax_issues
from app.core.db import get_session
from app.core.deps import current_user, require_op
from app.permissions.scope_filter import _is_super_admin
from app.table_tools import engine
from app.table_tools import ai_builder
from app.table_tools.dwd_relation_service import (
    apply_dwd_relation,
    list_dwd_fields,
    list_dwd_fields_by_dataset,
    list_dwd_sources,
    load_dataset_dwd_context,
    load_dwd_context,
    validate_relation_payload,
)
from app.table_tools.models import (
    MergeDwdRelation,
    MergeKeyMapping,
    MergePreviewRun,
    MergeResultBatch,
    MergeResultRow,
    MergeSourceMapping,
    MergeTemplate,
)
from app.users.models import User

router = APIRouter(prefix="/table-tools", tags=["table-tools"])

MENU = "table_tools"
AI_DRAFT_CAPABILITY = "table_merge.suggest_mapping"


# ── Schemas ────────────────────────────────────────────────
class SourceMappingIn(BaseModel):
    id: int | None = None
    name: str = Field(min_length=1, max_length=128)
    match_signature: list[str] = []
    source_fields: list[str] = []
    sheet_kw: str | None = None
    header_start: int = 1
    header_end: int = 1
    key_map: dict[str, str] = {}
    column_map: dict[str, str] = {}
    new_std_fields: list[str] = []
    derived_fields: list[dict] = []
    derive_check: dict | None = None
    skip_tokens: list[str] = ["合计", "小计", "总计"]


class TemplateIn(BaseModel):
    name: str = Field(min_length=1, max_length=128)
    description: str | None = None
    merge_keys: list[str] = Field(min_length=1)
    std_fields: list[str] = Field(min_length=1)
    aggregate: str = "sum"
    result_save_mode: str = "input_period"
    result_period_field: str | None = None
    mappings: list[SourceMappingIn] = []


class TemplateOut(BaseModel):
    id: int
    name: str
    description: str | None
    merge_keys: list[str]
    std_fields: list[str]
    aggregate: str
    result_save_mode: str
    result_period_field: str | None
    version: int
    mapping_count: int
    created_by: int | None


class MappingDraftMappingOut(SourceMappingIn):
    model_config = ConfigDict(populate_by_name=True)
    confidence: float = Field(default=0.0, validation_alias="_confidence", serialization_alias="_confidence")
    notes: str = Field(default="", validation_alias="_notes", serialization_alias="_notes")


class MappingDraftLowConfidenceOut(BaseModel):
    sheet: str
    confidence: float
    notes: str


class MappingDraftOut(BaseModel):
    mapping: MappingDraftMappingOut
    available_sheets: list[str]
    effective_headers: list[str]
    low_confidence: list[MappingDraftLowConfidenceOut]
    warnings: list[str]


class MappingDraftsOut(BaseModel):
    mappings: list[MappingDraftMappingOut]
    low_confidence: list[MappingDraftLowConfidenceOut]
    warnings: list[str]


class SourceMappingBatchIn(BaseModel):
    mappings: list[SourceMappingIn] = Field(min_length=1)


class KeyMappingIn(BaseModel):
    source_key: dict[str, Any] = Field(min_length=1)
    canonical_merge_key: dict[str, Any] = Field(min_length=1)
    enabled: bool = True


class DwdRelationIn(BaseModel):
    name: str = Field(min_length=1, max_length=128)
    dataset_id: int | None = None
    report_id: int | None = None
    left_fields: list[str] = Field(min_length=1)
    right_fields: list[str] = Field(min_length=1)
    select_fields: list[str] = []
    missing_policy: str = "anomaly"
    multiple_policy: str = "anomaly"
    enabled: bool = True


class SaveResultBatchIn(BaseModel):
    preview_token: str = Field(min_length=1, max_length=64)
    period: str | None = Field(default=None, min_length=6, max_length=6)


# ── 序列化 ─────────────────────────────────────────────────
def _mapping_to_engine(m: MergeSourceMapping) -> dict:
    return {
        "id": m.id,
        "name": m.name,
        "match": m.match_signature,
        "sheet_kw": m.sheet_kw,
        "header": [m.header_start, m.header_end],
        "key_map": m.key_map,
        "column_map": m.column_map,
        "derived_fields": m.derived_fields,
        "derive_check": m.derive_check,
        "skip_tokens": m.skip_tokens,
    }


def _template_out(t: MergeTemplate) -> TemplateOut:
    return TemplateOut(
        id=t.id, name=t.name, description=t.description,
        merge_keys=t.merge_keys, std_fields=t.std_fields,
        aggregate=t.aggregate, result_save_mode=t.result_save_mode,
        result_period_field=t.result_period_field, version=t.version,
        mapping_count=len(t.mappings), created_by=t.created_by,
    )


def _mapping_out(m: MergeSourceMapping) -> dict:
    return {
        "id": m.id,
        "name": m.name,
        "match_signature": m.match_signature,
        "source_fields": m.source_fields,
        "sheet_kw": m.sheet_kw,
        "header_start": m.header_start,
        "header_end": m.header_end,
        "key_map": m.key_map,
        "column_map": m.column_map,
        "derived_fields": m.derived_fields,
        "derive_check": m.derive_check,
        "skip_tokens": m.skip_tokens,
    }


def _validate_source_mapping(template: MergeTemplate, payload: SourceMappingIn) -> dict:
    name = payload.name.strip()
    if not name:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="映射名称不能为空")
    signature = list(dict.fromkeys(item.strip() for item in payload.match_signature if item.strip()))
    source_fields = list(dict.fromkeys(str(item).strip() for item in payload.source_fields if str(item).strip()))
    if not source_fields:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="请先上传样表解析源字段")
    if len(signature) < 3:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="表头特征至少需要 3 项")
    if any(field not in source_fields for field in signature):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="表头特征必须属于样表解析字段")
    if not 1 <= payload.header_start <= payload.header_end <= 10:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="表头行范围必须在 1 到 10 行之间")
    key_map = {key.strip(): value for key, value in payload.key_map.items() if key.strip()}
    if not key_map:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="至少配置一项主键映射")
    if any(value not in template.merge_keys for value in key_map.values()):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="主键映射目标必须属于模板归集主键")
    column_map = {key.strip(): value.strip() for key, value in payload.column_map.items() if key.strip()}
    if any(key not in source_fields for key in (*key_map.keys(), *column_map.keys())):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="映射源字段必须属于样表解析字段")
    new_std_fields = list(dict.fromkeys(str(field).strip() for field in (payload.new_std_fields or []) if str(field).strip()))
    allowed_std_fields = set(template.std_fields) | set(new_std_fields)
    if any(value not in allowed_std_fields for value in column_map.values()):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="字段映射目标必须属于模板标准字段或本次新增标准字段")
    derived_fields = payload.derived_fields or []
    targets: set[str] = set()
    for field in derived_fields:
        target = str(field.get("target") or "").strip()
        expr = str(field.get("expr") or "").strip()
        if target not in allowed_std_fields:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="派生字段目标必须属于模板标准字段或本次新增标准字段")
        if not expr:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="派生字段公式不能为空")
        if target in targets:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="同一映射不能重复配置派生字段目标")
        if target in column_map.values():
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="派生字段目标不能同时配置直接映射")
        refs = [item.strip() for item in re.findall(r"\{([^{}]+)\}", expr)]
        if not refs or any(ref not in allowed_std_fields for ref in refs):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="派生公式只能引用模板标准字段或本次新增标准字段")
        converted = re.sub(
            r"\{([^{}]+)\}", lambda match: f'FIELD({match.group(1).strip()!r})', expr
        )
        if formula_syntax_issues(converted):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="派生字段公式语法不合法")
        targets.add(target)
    derive_check = payload.derive_check
    if derive_check:
        sum_of = derive_check.get("sum_of") or []
        equals_col = str(derive_check.get("equals_col") or "").strip()
        if not sum_of or not equals_col or any(field not in allowed_std_fields for field in sum_of) or equals_col not in allowed_std_fields:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="拆分校验只能引用模板标准字段或本次新增标准字段")
        derive_check = {
            **derive_check,
            "sum_of": list(dict.fromkeys(sum_of)),
            "equals_col": equals_col,
        }
    derived_fields = [
        {**field, "target": str(field["target"]).strip(), "expr": str(field["expr"]).strip()}
        for field in derived_fields
    ]
    return {
        "name": name,
        "match_signature": signature,
        "source_fields": source_fields,
        "sheet_kw": payload.sheet_kw.strip() if payload.sheet_kw else None,
        "header_start": payload.header_start,
        "header_end": payload.header_end,
        "key_map": key_map,
        "column_map": column_map,
        "new_std_fields": new_std_fields,
        "derived_fields": derived_fields,
        "derive_check": derive_check,
        "skip_tokens": list(dict.fromkeys(item.strip() for item in (payload.skip_tokens or []) if item.strip())) or ["合计", "小计", "总计"],
    }


def _apply_source_mapping(mapping: MergeSourceMapping, values: dict) -> None:
    for key, value in values.items():
        setattr(mapping, key, value)


def _key_mapping_out(item: MergeKeyMapping) -> dict:
    return {
        "id": item.id,
        "template_id": item.template_id,
        "source_key": item.source_key,
        "canonical_merge_key": item.canonical_merge_key,
        "enabled": item.enabled,
    }


def _validate_key_mapping(template: MergeTemplate, payload: KeyMappingIn) -> dict:
    if set(payload.source_key) != set(template.merge_keys):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="源主键字段必须完整匹配模板归集主键")
    if set(payload.canonical_merge_key) != set(template.merge_keys):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="统一键字段必须完整匹配模板归集主键")
    return {
        "source_key": {field: payload.source_key[field] for field in template.merge_keys},
        "canonical_merge_key": {field: payload.canonical_merge_key[field] for field in template.merge_keys},
        "enabled": payload.enabled,
    }


async def _load_template(db: AsyncSession, tid: int) -> MergeTemplate:
    row = (await db.execute(
        select(MergeTemplate).where(MergeTemplate.id == tid)
        .options(
            selectinload(MergeTemplate.mappings),
            selectinload(MergeTemplate.key_mappings),
            selectinload(MergeTemplate.dwd_relations),
        )
    )).scalar_one_or_none()
    if row is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="模板不存在")
    return row


async def _ensure_can_modify(db: AsyncSession, t: MergeTemplate, user: User) -> None:
    """改/删门禁:仅模板创建者本人或超级管理员可操作。"""
    if t.created_by == user.id:
        return
    if await _is_super_admin(user, db):
        return
    raise HTTPException(
        status.HTTP_403_FORBIDDEN, detail="只能操作自己创建的模板"
    )


# ── 模板 CRUD ──────────────────────────────────────────────
@router.get("/templates", response_model=list[TemplateOut])
async def list_templates(
    db: AsyncSession = Depends(get_session),
    _: User = Depends(require_op(MENU, "V")),
) -> list[TemplateOut]:
    rows = (await db.execute(
        select(MergeTemplate).options(selectinload(MergeTemplate.mappings))
        .order_by(MergeTemplate.created_at.desc())
    )).scalars().all()
    return [_template_out(t) for t in rows]


@router.get("/templates/{tid}")
async def get_template(
    tid: int,
    db: AsyncSession = Depends(get_session),
    _: User = Depends(require_op(MENU, "V")),
) -> dict:
    t = await _load_template(db, tid)
    return {
        **_template_out(t).model_dump(),
        "mappings": [
            {
                "id": m.id,
                "name": m.name,
                "match_signature": m.match_signature,
                "source_fields": m.source_fields,
                "sheet_kw": m.sheet_kw,
                "header_start": m.header_start,
                "header_end": m.header_end,
                "key_map": m.key_map,
                "column_map": m.column_map,
                "derived_fields": m.derived_fields,
                "derive_check": m.derive_check,
                "skip_tokens": m.skip_tokens,
            }
            for m in t.mappings
        ],
        "key_mappings": [_key_mapping_out(item) for item in t.key_mappings],
        "dwd_relations": [
            {
                "id": item.id,
                "template_id": item.template_id,
                "name": item.name,
                "report_id": item.report_id,
                "dataset_id": item.dataset_id,
                "left_fields": item.left_fields,
                "right_fields": item.right_fields,
                "select_fields": item.select_fields,
                "missing_policy": item.missing_policy,
                "multiple_policy": item.multiple_policy,
                "enabled": item.enabled,
            }
            for item in t.dwd_relations
        ],
    }


@router.post("/templates", response_model=TemplateOut, status_code=status.HTTP_201_CREATED)
async def create_template(
    payload: TemplateIn,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "C")),
) -> TemplateOut:
    if (await db.execute(select(MergeTemplate).where(MergeTemplate.name == payload.name))).scalar_one_or_none():
        raise HTTPException(status.HTTP_409_CONFLICT, detail="模板名已存在")
    mode, period_field = _validate_save_config(payload.result_save_mode, payload.result_period_field, payload.std_fields)
    t = MergeTemplate(
        name=payload.name, description=payload.description,
        merge_keys=payload.merge_keys, std_fields=payload.std_fields,
        aggregate=payload.aggregate, result_save_mode=mode,
        result_period_field=period_field, created_by=user.id,
    )
    mapping_names: set[str] = set()
    for ms in payload.mappings:
        values = _validate_source_mapping(t, ms)
        new_fields = values.pop("new_std_fields", [])
        t.std_fields = list(dict.fromkeys([*t.std_fields, *new_fields]))
        if values["name"] in mapping_names:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="映射名称已存在")
        mapping_names.add(values["name"])
        t.mappings.append(MergeSourceMapping(**values))
    db.add(t)
    await db.commit()
    await db.refresh(t, ["mappings"])
    return _template_out(t)


@router.put("/templates/{tid}", response_model=TemplateOut)
async def update_template(
    tid: int,
    payload: TemplateIn,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "U")),
) -> TemplateOut:
    t = await _load_template(db, tid)
    await _ensure_can_modify(db, t, user)
    t.name = payload.name
    t.description = payload.description
    t.merge_keys = payload.merge_keys
    t.std_fields = payload.std_fields
    t.aggregate = payload.aggregate
    mode, period_field = _validate_save_config(payload.result_save_mode, payload.result_period_field, payload.std_fields)
    t.result_save_mode = mode
    t.result_period_field = period_field
    existing_mappings = {mapping.id: mapping for mapping in t.mappings}
    names: set[str] = set()
    mapping_ids: set[int] = set()
    prepared: list[tuple[int | None, dict]] = []
    for ms in payload.mappings:
        values = _validate_source_mapping(t, ms)
        if ms.id is not None:
            if ms.id in mapping_ids:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="映射 ID 不可重复")
            if ms.id not in existing_mappings:
                raise HTTPException(status.HTTP_404_NOT_FOUND, detail="源映射不存在")
            mapping_ids.add(ms.id)
        if values["name"] in names:
            raise HTTPException(status.HTTP_409_CONFLICT, detail="映射名称已存在")
        names.add(values["name"])
        prepared.append((ms.id, values))
    retained_ids = {mapping_id for mapping_id, _ in prepared if mapping_id is not None}
    for mapping_id, mapping in existing_mappings.items():
        if mapping_id not in retained_ids:
            await db.delete(mapping)
    for mapping_id, values in prepared:
        if mapping_id is None:
            t.mappings.append(MergeSourceMapping(**values))
        else:
            _apply_source_mapping(existing_mappings[mapping_id], values)
    t.version += 1
    await db.commit()
    await db.refresh(t, ["mappings"])
    return _template_out(t)


@router.post("/templates/{tid}/mappings/batch", status_code=status.HTTP_201_CREATED)
async def create_source_mappings_batch(
    tid: int,
    payload: SourceMappingBatchIn,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "U")),
) -> dict:
    """校验全部源映射后原子新增，避免批量上传产生部分写入。"""
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    names = {item.name for item in template.mappings}
    prepared: list[dict] = []
    for mapping in payload.mappings:
        values = _validate_source_mapping(template, mapping)
        new_fields = values.pop("new_std_fields", [])
        template.std_fields = list(dict.fromkeys([*template.std_fields, *new_fields]))
        if values["name"] in names:
            raise HTTPException(status.HTTP_409_CONFLICT, detail=f"映射名称已存在: {values['name']}")
        names.add(values["name"])
        prepared.append(values)

    created = [MergeSourceMapping(**values) for values in prepared]
    template.mappings.extend(created)
    template.version += 1
    await db.commit()
    for mapping in created:
        await db.refresh(mapping)
    return {"mappings": [_mapping_out(mapping) for mapping in created]}

# ── 单条源映射维护 ──────────────────────────────────────────
@router.post("/templates/{tid}/mappings", status_code=status.HTTP_201_CREATED)
async def create_source_mapping(tid: int, payload: SourceMappingIn, db: AsyncSession = Depends(get_session), user: User = Depends(require_op(MENU, "U"))) -> dict:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    values = _validate_source_mapping(template, payload)
    new_fields = values.pop("new_std_fields", [])
    template.std_fields = list(dict.fromkeys([*template.std_fields, *new_fields]))
    if any(item.name == values["name"] for item in template.mappings):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="映射名称已存在")
    mapping = MergeSourceMapping(**values)
    template.mappings.append(mapping)
    template.version += 1
    await db.commit()
    await db.refresh(mapping)
    return _mapping_out(mapping)


@router.put("/templates/{tid}/mappings/{mid}")
async def update_source_mapping(tid: int, mid: int, payload: SourceMappingIn, db: AsyncSession = Depends(get_session), user: User = Depends(require_op(MENU, "U"))) -> dict:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    mapping = next((item for item in template.mappings if item.id == mid), None)
    if mapping is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="源映射不存在")
    values = _validate_source_mapping(template, payload)
    if any(item.id != mid and item.name == values["name"] for item in template.mappings):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="映射名称已存在")
    _apply_source_mapping(mapping, values)
    template.version += 1
    await db.commit()
    await db.refresh(mapping)
    return _mapping_out(mapping)


@router.delete("/templates/{tid}/mappings/{mid}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_source_mapping(tid: int, mid: int, db: AsyncSession = Depends(get_session), user: User = Depends(require_op(MENU, "D"))) -> Response:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    mapping = next((item for item in template.mappings if item.id == mid), None)
    if mapping is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="源映射不存在")
    await db.delete(mapping)
    template.version += 1
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)

# ── 主键值映射维护 ────────────────────────────────────────────
@router.get("/templates/{tid}/key-mappings")
async def list_key_mappings(
    tid: int,
    db: AsyncSession = Depends(get_session),
    _: User = Depends(require_op(MENU, "V")),
) -> list[dict]:
    template = await _load_template(db, tid)
    return [_key_mapping_out(item) for item in template.key_mappings]


@router.post("/templates/{tid}/key-mappings", status_code=status.HTTP_201_CREATED)
async def create_key_mapping(
    tid: int,
    payload: KeyMappingIn,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "U")),
) -> dict:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    values = _validate_key_mapping(template, payload)
    if any(
        item.source_key == values["source_key"]
        for item in template.key_mappings
    ):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="源主键映射已存在")
    item = MergeKeyMapping(template_id=tid, created_by=user.id, **values)
    template.key_mappings.append(item)
    template.version += 1
    await db.commit()
    await db.refresh(item)
    return _key_mapping_out(item)


@router.put("/templates/{tid}/key-mappings/{mid}")
async def update_key_mapping(
    tid: int,
    mid: int,
    payload: KeyMappingIn,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "U")),
) -> dict:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    item = next((value for value in template.key_mappings if value.id == mid), None)
    if item is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="主键映射不存在")
    values = _validate_key_mapping(template, payload)
    if any(
        value.id != mid
        and value.source_key == values["source_key"]
        for value in template.key_mappings
    ):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="源主键映射已存在")
    for key, value in values.items():
        setattr(item, key, value)
    template.version += 1
    await db.commit()
    await db.refresh(item)
    return _key_mapping_out(item)


@router.delete("/templates/{tid}/key-mappings/{mid}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_key_mapping(
    tid: int,
    mid: int,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "D")),
) -> Response:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    item = next((value for value in template.key_mappings if value.id == mid), None)
    if item is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="主键映射不存在")
    await db.delete(item)
    template.version += 1
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


async def delete_template(
    tid: int,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "D")),
) -> Response:
    t = await _load_template(db, tid)
    await _ensure_can_modify(db, t, user)
    await db.delete(t)
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ── DWD 来源与独立关联配置 ────────────────────────────────────
@router.get("/dwd-relation-sources")
async def dwd_relation_sources(
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "V")),
) -> list[dict]:
    return await list_dwd_sources(user, db)


@router.get("/datasets/{dataset_id}/dwd-fields")
async def dwd_dataset_fields(
    dataset_id: int,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "V")),
) -> list[dict]:
    return await list_dwd_fields_by_dataset(dataset_id, user, db)


@router.get("/reports/{report_id}/dwd-fields")
async def dwd_fields(
    report_id: int,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "V")),
) -> list[dict]:
    return await list_dwd_fields(report_id, user, db)


@router.get("/templates/{tid}/dwd-relations")
async def list_dwd_relations(
    tid: int,
    db: AsyncSession = Depends(get_session),
    _: User = Depends(require_op(MENU, "V")),
) -> list[dict]:
    template = await _load_template(db, tid)
    return [{
        "id": item.id,
        "template_id": item.template_id,
        "name": item.name,
        "report_id": item.report_id,
        "dataset_id": item.dataset_id,
        "left_fields": item.left_fields,
        "right_fields": item.right_fields,
        "select_fields": item.select_fields,
        "missing_policy": item.missing_policy,
        "multiple_policy": item.multiple_policy,
        "enabled": item.enabled,
    } for item in template.dwd_relations]


@router.post("/templates/{tid}/dwd-relations", status_code=status.HTTP_201_CREATED)
async def create_dwd_relation(
    tid: int,
    payload: DwdRelationIn,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "U")),
) -> dict:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    if payload.dataset_id is None and payload.report_id is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="必须指定 DWD 数据集")
    values = validate_relation_payload(
        payload.model_dump(), template.merge_keys + template.std_fields,
        [
            item["code"]
            for item in await (
                list_dwd_fields_by_dataset(payload.dataset_id, user, db)
                if payload.dataset_id is not None
                else list_dwd_fields(payload.report_id, user, db)
            )
        ],
    )
    if values["dataset_id"] is not None:
        await load_dataset_dwd_context(values["dataset_id"], user, db)
        values["report_id"] = None
    else:
        await load_dwd_context(values["report_id"], user, db)
    if any(item.name == values["name"] for item in template.dwd_relations):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="DWD 关联名称已存在")
    item = MergeDwdRelation(template_id=tid, created_by=user.id, **values)
    template.dwd_relations.append(item)
    template.version += 1
    await db.commit()
    await db.refresh(item)
    return {"id": item.id, "template_id": tid, **values}


@router.put("/templates/{tid}/dwd-relations/{rid}")
async def update_dwd_relation(
    tid: int,
    rid: int,
    payload: DwdRelationIn,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "U")),
) -> dict:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    item = next((value for value in template.dwd_relations if value.id == rid), None)
    if item is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="DWD 关联不存在")
    if payload.dataset_id is None and payload.report_id is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="必须指定 DWD 数据集")
    values = validate_relation_payload(
        payload.model_dump(), template.merge_keys + template.std_fields,
        [
            item["code"]
            for item in await (
                list_dwd_fields_by_dataset(payload.dataset_id, user, db)
                if payload.dataset_id is not None
                else list_dwd_fields(payload.report_id, user, db)
            )
        ],
    )
    if values["dataset_id"] is not None:
        await load_dataset_dwd_context(values["dataset_id"], user, db)
        values["report_id"] = None
    else:
        await load_dwd_context(values["report_id"], user, db)
    if any(value.id != rid and value.name == values["name"] for value in template.dwd_relations):
        raise HTTPException(status.HTTP_409_CONFLICT, detail="DWD 关联名称已存在")
    for key, value in values.items():
        setattr(item, key, value)
    template.version += 1
    await db.commit()
    await db.refresh(item)
    return {"id": item.id, "template_id": tid, **values}


@router.delete("/templates/{tid}/dwd-relations/{rid}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_dwd_relation(
    tid: int,
    rid: int,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "D")),
) -> Response:
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    item = next((value for value in template.dwd_relations if value.id == rid), None)
    if item is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="DWD 关联不存在")
    await db.delete(item)
    template.version += 1
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


async def _run_template_merge(
    template: MergeTemplate,
    blobs: list[tuple[str, bytes]],
    user: User,
    db: AsyncSession,
    relations: list[MergeDwdRelation] | None = None,
) -> dict:
    template_config = {
        "merge_keys": template.merge_keys,
        "std_fields": template.std_fields,
        "aggregate": template.aggregate,
    }
    mappings = [_mapping_to_engine(item) for item in template.mappings]
    key_mappings = [
        {
            "source_key": item.source_key,
            "canonical_merge_key": item.canonical_merge_key,
        }
        for item in template.key_mappings if item.enabled
    ]
    try:
        result = engine.run_merge(
            blobs, template_config, mappings, await executable_functions(db), key_mappings
        )
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    dwd_anomalies: list[dict[str, Any]] = []
    rows = result["rows"]
    base_columns = [column for column in result["columns"] if column != "来源"]
    dwd_columns: list[str] = []
    column_labels: dict[str, str] = {}
    for relation in relations if relations is not None else template.dwd_relations:
        if not relation.enabled:
            continue
        rows, anomalies = await apply_dwd_relation(rows, relation, user, db)
        dwd_anomalies.extend(anomalies)
        fields = await (
            list_dwd_fields_by_dataset(relation.dataset_id, user, db)
            if relation.dataset_id is not None
            else list_dwd_fields(relation.report_id, user, db)
        )
        field_labels = {field["code"]: field["label"] for field in fields}
        for field in relation.select_fields:
            if field not in dwd_columns:
                dwd_columns.append(field)
            if field in field_labels:
                column_labels[field] = field_labels[field]
    columns = base_columns + dwd_columns + (["来源"] if "来源" in result["columns"] else [])
    return {
        **result,
        "rows": rows,
        "columns": columns,
        "column_labels": column_labels,
        "dwd_anomalies": dwd_anomalies,
    }


@router.post("/templates/{tid}/dwd-relations/{rid}/apply")
async def apply_dwd_relation_api(
    tid: int,
    rid: int,
    files: list[UploadFile] = File(...),
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "V")),
) -> dict:
    template = await _load_template(db, tid)
    relation = next((value for value in template.dwd_relations if value.id == rid), None)
    if relation is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="DWD 关联不存在")
    blobs = await _read_files(files)
    result = await _run_template_merge(template, blobs, user, db, relations=[relation])
    return {**result, "rows": result["rows"][:100], "total_rows": len(result["rows"])}


async def _read_files(files: list[UploadFile]) -> list[tuple[str, bytes]]:
    out: list[tuple[str, bytes]] = []
    for f in files:
        name = f.filename or ""
        if name.startswith(("_", "~$")) or not name.lower().endswith((".xlsx", ".xls")):
            continue
        content = await f.read()
        if content:
            out.append((name, content))
    if not out:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="未收到有效的 Excel 文件")
    return out


def _validate_period(period: str) -> str:
    if len(period) != 6 or not period.isdigit() or not 1 <= int(period[4:]) <= 12:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="业务月份必须为有效的 YYYYMM 格式")
    return period


def _validate_save_config(mode: str, period_field: str | None, std_fields: list[str]) -> tuple[str, str | None]:
    if mode not in {"none", "input_period", "field_period"}:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="结果保存方式无效")
    if mode == "field_period":
        if not period_field or period_field not in std_fields:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="从结果字段读取期间时，期间字段必须属于标准字段")
        return mode, period_field
    if period_field:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="当前保存方式不能设置业务期间字段")
    return mode, None


def _row_key(row: dict[str, Any], merge_keys: list[str]) -> tuple[dict[str, str], str]:
    values = {field: str(row.get(field, "") or "").strip() for field in merge_keys}
    if any(not value for value in values.values()):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="预览结果存在不完整的归集主键，不能保存")
    encoded = json.dumps(values, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return values, hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _batch_out(batch: MergeResultBatch) -> dict[str, Any]:
    return {
        "id": batch.id,
        "period": batch.period,
        "template_version": batch.template_version,
        "row_count": batch.row_count,
        "created_at": batch.created_at,
        "updated_at": batch.updated_at,
    }


@router.post("/templates/{tid}/merge")
async def run_merge_api(
    tid: int,
    files: list[UploadFile] = File(...),
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "V")),
) -> dict:
    """用模板跑合并并持久化完整预览，响应只返回前 100 行。"""
    template = await _load_template(db, tid)
    result = await _run_template_merge(template, await _read_files(files), user=user, db=db)
    preview = MergePreviewRun(
        token=secrets.token_urlsafe(32),
        template_id=template.id,
        template_version=template.version,
        merge_keys_snapshot=list(template.merge_keys),
        columns_snapshot=list(result["columns"]),
        rows=result["rows"],
        stats=result["stats"],
        recognize_log=result["recognize_log"],
        anomalies=result["anomalies"],
        dwd_anomalies=result["dwd_anomalies"],
        created_by=user.id,
        expires_at=datetime.now(timezone.utc) + timedelta(hours=2),
    )
    db.add(preview)
    await db.commit()
    return {
        "preview_token": preview.token,
        "columns": result["columns"],
        "column_labels": result["column_labels"],
        "rows": result["rows"][:100],
        "total_rows": len(result["rows"]),
        "recognize_log": result["recognize_log"],
        "anomalies": result["anomalies"],
        "dwd_anomalies": result["dwd_anomalies"],
        "stats": result["stats"],
        "key_mapping_stats": result["key_mapping_stats"],
        "raw_key_traces": result["raw_key_traces"],
    }


@router.post("/templates/{tid}/result-batches/save")
async def save_result_batch(
    tid: int,
    payload: SaveResultBatchIn,
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "V")),
) -> dict:
    template = await _load_template(db, tid)
    mode, period_field = _validate_save_config(template.result_save_mode, template.result_period_field, template.std_fields)
    preview = (await db.execute(
        select(MergePreviewRun).where(
            MergePreviewRun.token == payload.preview_token,
            MergePreviewRun.template_id == tid,
            MergePreviewRun.created_by == user.id,
        )
    )).scalar_one_or_none()
    if preview is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="预览结果不存在或无权保存")
    if preview.expires_at < datetime.now(timezone.utc):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="预览结果已过期，请重新运行预览")

    if mode == "none":
        period = "CURRENT"
    elif mode == "input_period":
        if payload.period is None:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="请输入业务月份 YYYYMM")
        period = _validate_period(payload.period)
    else:
        if payload.period is not None:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="当前模板的业务月份必须从结果字段读取")
        if period_field not in preview.columns_snapshot:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="预览结果不包含配置的业务月份字段")
        if not preview.rows:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="没有可保存的结果，无法识别业务月份")
        periods = set()
        for row in preview.rows:
            value = str(row.get(period_field or "", "") or "").strip()
            if not value:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="结果中存在缺失的业务月份")
            periods.add(_validate_period(value))
        if len(periods) != 1:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="结果中必须包含唯一的有效业务月份")
        period = periods.pop()

    keyed_rows: list[tuple[dict[str, str], str, dict[str, Any]]] = []
    hashes: set[str] = set()
    for row in preview.rows:
        key, key_hash = _row_key(row, preview.merge_keys_snapshot)
        if key_hash in hashes:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="预览结果存在重复的归集联合主键，不能保存")
        hashes.add(key_hash)
        keyed_rows.append((key, key_hash, row))

    batch = (await db.execute(
        select(MergeResultBatch).where(
            MergeResultBatch.template_id == tid, MergeResultBatch.period == period
        ).with_for_update()
    )).scalar_one_or_none()
    if batch is None:
        batch = MergeResultBatch(
            template_id=tid, period=period, template_version=preview.template_version,
            merge_keys_snapshot=preview.merge_keys_snapshot, columns_snapshot=preview.columns_snapshot,
            stats=preview.stats, anomalies=preview.anomalies, dwd_anomalies=preview.dwd_anomalies,
            created_by=user.id, updated_by=user.id,
        )
        db.add(batch)
        await db.flush()
        existing: dict[str, MergeResultRow] = {}
    else:
        existing_rows = (await db.execute(
            select(MergeResultRow).where(MergeResultRow.batch_id == batch.id)
        )).scalars().all()
        existing = {item.merge_key_hash: item for item in existing_rows}
        batch.template_version = preview.template_version
        batch.merge_keys_snapshot = preview.merge_keys_snapshot
        batch.columns_snapshot = preview.columns_snapshot
        batch.stats = preview.stats
        batch.anomalies = preview.anomalies
        batch.dwd_anomalies = preview.dwd_anomalies
        batch.updated_by = user.id

    inserted_count = 0
    replaced_count = 0
    for key, key_hash, values in keyed_rows:
        current = existing.get(key_hash)
        if current is None:
            db.add(MergeResultRow(batch_id=batch.id, merge_key=key, merge_key_hash=key_hash, values=values))
            inserted_count += 1
        else:
            if current.merge_key != key:
                raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="归集主键校验冲突，不能保存")
            current.values = values
            replaced_count += 1
    await db.flush()
    batch.row_count = len(existing) + inserted_count
    await db.commit()
    return {**_batch_out(batch), "inserted_count": inserted_count, "replaced_count": replaced_count, "total_count": batch.row_count}


@router.get("/templates/{tid}/result-batches")
async def list_result_batches(
    tid: int,
    db: AsyncSession = Depends(get_session),
    _: User = Depends(require_op(MENU, "V")),
) -> list[dict]:
    await _load_template(db, tid)
    batches = (await db.execute(
        select(MergeResultBatch).where(MergeResultBatch.template_id == tid).order_by(MergeResultBatch.period.desc())
    )).scalars().all()
    return [_batch_out(batch) for batch in batches]


@router.get("/templates/{tid}/result-batches/{bid}/rows")
async def list_result_batch_rows(
    tid: int, bid: int, page: int = 1, page_size: int = 100,
    db: AsyncSession = Depends(get_session),
    _: User = Depends(require_op(MENU, "V")),
) -> dict:
    if page < 1 or not 1 <= page_size <= 500:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="分页参数无效")
    batch = (await db.execute(select(MergeResultBatch).where(MergeResultBatch.id == bid, MergeResultBatch.template_id == tid))).scalar_one_or_none()
    if batch is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="历史结果不存在")
    rows = (await db.execute(
        select(MergeResultRow).where(MergeResultRow.batch_id == bid).order_by(MergeResultRow.id).offset((page - 1) * page_size).limit(page_size)
    )).scalars().all()
    return {"batch": _batch_out(batch), "columns": batch.columns_snapshot, "rows": [row.values for row in rows], "total_rows": batch.row_count}


@router.get("/templates/{tid}/result-batches/{bid}/download")
async def download_result_batch(
    tid: int, bid: int,
    db: AsyncSession = Depends(get_session),
    _: User = Depends(require_op(MENU, "E")),
) -> Response:
    batch = (await db.execute(select(MergeResultBatch).where(MergeResultBatch.id == bid, MergeResultBatch.template_id == tid))).scalar_one_or_none()
    if batch is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="历史结果不存在")
    rows = (await db.execute(select(MergeResultRow).where(MergeResultRow.batch_id == bid).order_by(MergeResultRow.id))).scalars().all()
    xlsx = engine.rows_to_xlsx(batch.columns_snapshot, [row.values for row in rows])
    return Response(content=xlsx, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="merged_result_{batch.period}.xlsx"'})


@router.post("/templates/{tid}/download")
async def download_merge(
    tid: int,
    files: list[UploadFile] = File(...),
    db: AsyncSession = Depends(get_session),
    _: User = Depends(require_op(MENU, "E")),
) -> Response:
    """跑合并并直接下载 xlsx(导出口径 E)。"""
    t = await _load_template(db, tid)
    blobs = await _read_files(files)
    result = await _run_template_merge(t, blobs, user=_, db=db)
    xlsx = engine.rows_to_xlsx(result["columns"], result["rows"])
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="merged_result.xlsx"'},
    )


@router.post("/templates/{tid}/mapping-drafts", response_model=MappingDraftsOut)
async def mapping_drafts(
    tid: int,
    files: list[UploadFile] = File(...),
    business_context: str = "",
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "U")),
) -> MappingDraftsOut:
    """批量上传样表，为既有模板生成待确认的源映射草稿。"""
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    capability = get_capability(AI_DRAFT_CAPABILITY)
    if capability is None:
        raise HTTPException(status.HTTP_503_SERVICE_UNAVAILABLE, detail="AI 能力未注册")
    try:
        validate_capability_policy(capability)
    except Exception as exc:
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail=f"AI 能力未启用: {exc}") from exc

    blobs = await _read_files(files)
    timer = AiAuditTimer()
    timer.add_event("entry", capability_id=AI_DRAFT_CAPABILITY)
    status_text = "ok"
    error: str | None = None
    mappings: list[dict] = []
    try:
        timer.add_event("model_call", capability_id=AI_DRAFT_CAPABILITY)
        mappings = await ai_builder.build_mapping_drafts(
            blobs, template.std_fields, template.merge_keys, business_context, db,
        )
        for mapping in mappings:
            source_fields = list(dict.fromkeys(
                str(field).strip() for field in (mapping.get("source_fields") or []) if str(field).strip()
            ))
            if not source_fields:
                raise ValueError("未从样表解析到源字段")
            mapping["source_fields"] = source_fields
        enforce_output_deny_patterns(capability, _draft_scan_text({"std_fields": [], "mappings": mappings}))
    except Exception as exc:
        status_text = "error"
        error = str(exc)

    await record_ai_log(
        db=db,
        user=user,
        action="table_merge_suggest_mapping",
        request_summary=f"批量单映射草稿 {len(blobs)} 个文件 {business_context[:60]}",
        response_summary=f"生成 {len(mappings)} 条映射草稿" if status_text == "ok" else (error or ""),
        input_payload={"files": [name for name, _ in blobs], "business_context": business_context},
        output_payload={"mapping": _draft_scan_text({"std_fields": [], "mappings": mappings})},
        status=status_text,
        error=error,
        metadata={"capability_id": AI_DRAFT_CAPABILITY, "metadata_only": True},
        timer=timer,
    )
    await db.commit()
    if status_text != "ok":
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail=error or "AI 生成失败")

    low_confidence = [
        {"sheet": mapping["name"], "confidence": float(mapping.get("_confidence", 0.0)), "notes": mapping.get("_notes", "")}
        for mapping in mappings
        if float(mapping.get("_confidence", 0.0)) < 0.85
    ]
    return MappingDraftsOut(mappings=mappings, low_confidence=low_confidence, warnings=[])

@router.post("/templates/{tid}/mapping-draft", response_model=MappingDraftOut)
async def mapping_draft(
    tid: int,
    file: UploadFile = File(...),
    sheet_name: str | None = None,
    header_start: int = 1,
    header_end: int = 1,
    business_context: str = "",
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "U")),
) -> dict:
    """只发送有效表头与既有模板字段，生成一条待人工确认的 AI 映射草稿。"""
    template = await _load_template(db, tid)
    await _ensure_can_modify(db, template, user)
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="请上传 .xlsx Excel 文件")
    if not 1 <= header_start <= header_end <= 10:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="表头行范围必须在 1 到 10 行之间")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="样表不得超过 10MB")
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Excel 解析失败: {exc}") from exc
    try:
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="指定 Sheet 不存在")
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        headers = [item for item in engine.parse_header(worksheet, header_start, header_end) if item]
        if not headers:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="未解析到有效表头")
        fallback = {
            "name": f"{filename.rsplit('.', 1)[0]}-映射", "match_signature": headers[:3],
            "sheet_kw": worksheet.title, "header_start": header_start, "header_end": header_end,
            "key_map": {}, "column_map": {}, "derived_fields": [], "derive_check": None,
            "source_fields": headers,
            "skip_tokens": ["合计", "小计", "总计"], "_confidence": 0.0,
            "_notes": "AI 不可用，已生成表头草稿，请手工配置映射。",
        }
        timer = AiAuditTimer()
        timer.add_event("entry", capability_id=AI_DRAFT_CAPABILITY)
        capability = get_capability(AI_DRAFT_CAPABILITY)
        warning: str | None = None
        mapping = fallback
        status_text = "fallback"
        if capability is None:
            warning = "AI 能力未注册，已切换为手工草稿"
        else:
            try:
                validate_capability_policy(capability)
                timer.add_event("model_call", capability_id=AI_DRAFT_CAPABILITY)
                mapping = await ai_builder.build_mapping_draft(
                    {"file": filename, "sheet": worksheet.title, "columns": headers,
                     "header_start": header_start, "header_end": header_end},
                    template.std_fields, template.merge_keys, business_context, db,
                )
                mapping["source_fields"] = headers
                enforce_output_deny_patterns(capability, _draft_scan_text({"std_fields": [], "mappings": [mapping]}))
                status_text = "ok"
            except Exception as exc:
                warning = f"AI 建议不可用，已切换为手工草稿：{exc}"
        confidence = float(mapping.get("_confidence", 0.0))
        await record_ai_log(
            db=db, user=user, action="table_merge_suggest_mapping",
            request_summary=f"单映射草稿 {filename}/{worksheet.title}",
            response_summary=f"置信度 {confidence:.2f}" if status_text == "ok" else (warning or ""),
            input_payload={"file": filename, "sheet": worksheet.title, "headers": headers, "business_context": business_context},
            output_payload={"mapping": _draft_scan_text({"std_fields": [], "mappings": [mapping]})},
            status=status_text, error=warning, metadata={"capability_id": AI_DRAFT_CAPABILITY, "metadata_only": True}, timer=timer,
        )
        await db.commit()
        return {
            "mapping": mapping, "available_sheets": workbook.sheetnames, "effective_headers": headers,
            "low_confidence": ([{"sheet": worksheet.title, "confidence": confidence, "notes": mapping.get("_notes", "")}]
                               if confidence < 0.85 else []),
            "warnings": [warning] if warning else [],
        }
    finally:
        workbook.close()
# ── AI 建模板草稿(走 004 底座:capability 注册 + 策略闸门 + 输出 deny + 审计)──────

@router.post("/ai-draft")
async def ai_draft(
    files: list[UploadFile] = File(...),
    business_context: str = "",
    db: AsyncSession = Depends(get_session),
    user: User = Depends(require_op(MENU, "E")),
) -> dict:
    """上传文件 + 可选业务背景 → AI 生成归集模板草稿（不存库）。

    AI 接入受 004 底座管控:能力注册表 + 策略闸门 + 输出 deny 扫描 + 统一审计。
    只发表头给模型(ai_builder 内只解析表头列名),明细行不进上下文(§4.8)。
    返回结构兼容 TemplateIn,附带 _meta.low_confidence 供前端标红。前端确认后 POST /templates 存库。
    """
    timer = AiAuditTimer()
    timer.add_event("entry", capability_id=AI_DRAFT_CAPABILITY)

    capability = get_capability(AI_DRAFT_CAPABILITY)
    if capability is None:
        raise HTTPException(status.HTTP_503_SERVICE_UNAVAILABLE, detail="AI 能力未注册")
    try:
        validate_capability_policy(capability)
    except Exception as e:
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail=f"AI 能力未启用: {e}")

    blobs = await _read_files(files)
    status_text = "ok"
    error: str | None = None
    draft: dict[str, Any] = {}
    try:
        timer.add_event("model_call", capability_id=AI_DRAFT_CAPABILITY)
        draft = await ai_builder.build_draft(blobs, business_context, db)
        # 输出级 deny:扫描 AI 产出的标准字段/映射文本,拦截 SQL/代码/URL 等注入
        enforce_output_deny_patterns(capability, _draft_scan_text(draft))
    except ValueError as e:
        status_text = "error"
        error = str(e)
    except Exception as e:  # policy deny 等
        status_text = "error"
        error = str(e)

    await record_ai_log(
        db=db,
        user=user,
        action="table_merge_suggest_mapping",
        request_summary=f"{len(blobs)}个文件 {business_context[:60]}",
        response_summary=f"{len(draft.get('std_fields', []))}个标准字段/{len(draft.get('mappings', []))}个源映射" if status_text == "ok" else (error or ""),
        input_payload={"files": [n for n, _ in blobs], "business_context": business_context},
        output_payload={"std_fields": draft.get("std_fields", []), "meta": draft.get("_meta", {})},
        status=status_text,
        error=error,
        metadata={"capability_id": AI_DRAFT_CAPABILITY},
        timer=timer,
    )
    await db.commit()

    if status_text != "ok":
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail=error or "AI 生成失败")
    return draft


def _draft_scan_text(draft: dict[str, Any]) -> str:
    """把草稿里模型生成的文本(标准字段名 + 各源 column_map/derived 表达式)拼起来供 deny 扫描。"""
    parts: list[str] = list(draft.get("std_fields") or [])
    for m in draft.get("mappings") or []:
        parts.extend((m.get("column_map") or {}).keys())
        parts.extend(str(v) for v in (m.get("column_map") or {}).values())
        for d in m.get("derived_fields") or []:
            parts.append(str(d.get("expr", "")))
            parts.append(str(d.get("target", "")))
    return "\n".join(parts)
