"""UCP Excel 文件导入服务 (Phase 2-7)

支持 .xlsx / .xls 文件导入为数据源：
  1. 上传文件 → 解析表头 + 预览脱敏样本 → 返回 file_key
  2. 导入到目标表 → 按映射规则转换 + upsert + 错误行记录

设计要点：
  - 文件临时存储在 UCP_EXCEL_UPLOAD_DIR（默认 backend/.tmp/ucp_excel）
  - 预览数据自动脱敏（薪酬/手机号/身份证/银行卡）
  - 导入失败行单独记录到返回结果，不阻断整批
  - EXCEL_IMPORT_ADAPTER 复用 parse_excel_file，可在 Pipeline 步骤中使用
"""
from __future__ import annotations

import logging
import os
import uuid
from datetime import datetime, UTC
from pathlib import Path
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.ucp.masking import mask_sensitive_fields

logger = logging.getLogger("ucp.excel_service")

# 临时文件目录
_DEFAULT_UPLOAD_DIR = Path(__file__).resolve().parents[3] / ".tmp" / "ucp_excel"
UPLOAD_DIR = Path(os.environ.get("UCP_EXCEL_UPLOAD_DIR", str(_DEFAULT_UPLOAD_DIR)))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# 允许的扩展名
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
# 单文件最大 20MB
MAX_FILE_SIZE = 20 * 1024 * 1024
# 预览行数
PREVIEW_ROW_LIMIT = 20


class ExcelImportError(Exception):
    """Excel 导入业务错误。"""
    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(f"[{code}] {message}")


def _validate_extension(filename: str) -> str:
    """校验文件扩展名，返回小写扩展名。"""
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ExcelImportError(
            "UNSUPPORTED_FILE_TYPE",
            f"不支持的文件类型 '{ext}'，仅支持 {', '.join(ALLOWED_EXTENSIONS)}",
        )
    return ext


def parse_excel_file(file_path: str | Path, sheet_name: str | None = None) -> dict:
    """解析 Excel 文件，返回 {headers, rows, total_rows}。

    第一行作为表头。空行自动跳过。
    """
    import openpyxl

    file_path = Path(file_path)
    if not file_path.exists():
        raise ExcelImportError("FILE_NOT_FOUND", f"文件不存在: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        raise ExcelImportError("PARSE_FAILED", f"Excel 解析失败: {str(e)[:300]}")

    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ExcelImportError("EMPTY_FILE", "文件为空（无表头行）")

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        if not any(headers):
            raise ExcelImportError("EMPTY_HEADER", "表头行为空")

        data_rows: list[dict] = []
        for row in rows_iter:
            # 跳过全空行
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            record = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    record[headers[idx]] = _normalize_cell(value)
            data_rows.append(record)
    finally:
        wb.close()

    return {
        "headers": headers,
        "rows": data_rows,
        "total_rows": len(data_rows),
    }


def _normalize_cell(value: Any) -> Any:
    """规范化单元格值：datetime 保留，其余转字符串（去除首尾空白）。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float, bool)):
        return value
    return str(value).strip()


async def save_and_preview(
    file_bytes: bytes,
    filename: str,
    sheet_name: str | None = None,
) -> dict:
    """保存上传文件并返回预览（脱敏样本）+ file_key。

    Returns:
        {
          file_key: str,           # 后续导入用
          filename: str,
          sheet_names: list[str],
          headers: list[str],
          preview_rows: list[dict],  # 脱敏后，最多 PREVIEW_ROW_LIMIT 行
          total_rows: int,
        }
    """
    ext = _validate_extension(filename)
    if len(file_bytes) > MAX_FILE_SIZE:
        raise ExcelImportError(
            "FILE_TOO_LARGE",
            f"文件超过 {MAX_FILE_SIZE // 1024 // 1024}MB 限制",
        )

    file_key = f"excel_{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = UPLOAD_DIR / file_key
    file_path.write_bytes(file_bytes)
    logger.info("[ucp.excel] saved upload: %s (%d bytes)", file_key, len(file_bytes))

    parsed = parse_excel_file(file_path, sheet_name=sheet_name)

    # 重新打开获取 sheet 列表
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()
    except Exception:
        sheet_names = []

    # 预览脱敏
    preview = parsed["rows"][:PREVIEW_ROW_LIMIT]
    masked_preview = mask_sensitive_fields(preview) if preview else []

    return {
        "file_key": file_key,
        "filename": filename,
        "sheet_names": sheet_names,
        "headers": parsed["headers"],
        "preview_rows": masked_preview,
        "total_rows": parsed["total_rows"],
    }


def resolve_file_path(file_key: str) -> Path:
    """根据 file_key 解析文件绝对路径，校验防目录穿越。"""
    if not file_key or "/" in file_key or "\\" in file_key:
        raise ExcelImportError("INVALID_FILE_KEY", f"非法 file_key: {file_key}")
    path = (UPLOAD_DIR / file_key).resolve()
    if not str(path).startswith(str(UPLOAD_DIR.resolve())):
        raise ExcelImportError("INVALID_FILE_KEY", f"非法 file_key: {file_key}")
    if not path.exists():
        raise ExcelImportError("FILE_NOT_FOUND", f"文件不存在或已过期: {file_key}")
    return path


async def import_to_target_table(
    db: AsyncSession,
    file_key: str,
    target_table: str,
    join_key: str,
    mapping_rules: list[dict] | None = None,
    sheet_name: str | None = None,
) -> dict:
    """将 Excel 数据导入目标表（按 mapping_rules 映射 + join_key upsert）。

    Returns:
        {
          status: SUCCESS / PARTIAL_SUCCESS / FAILED,
          total_rows: int,
          success_count: int,
          failed_count: int,
          failed_details: list[{row_index, reason}],
          target_table: str,
        }
    """
    from app.ucp.upsert_service import upsert_to_target_table

    file_path = resolve_file_path(file_key)
    parsed = parse_excel_file(file_path, sheet_name=sheet_name)
    rows = parsed["rows"]

    if not rows:
        return {
            "status": "SUCCESS",
            "total_rows": 0,
            "success_count": 0,
            "failed_count": 0,
            "failed_details": [],
            "target_table": target_table,
        }

    if not target_table:
        raise ExcelImportError("MISSING_TARGET_TABLE", "未指定目标表")
    if not join_key:
        raise ExcelImportError("MISSING_JOIN_KEY", "未指定幂等主键 join_key")

    # 应用字段映射
    mapped_rows: list[dict] = []
    failed_details: list[dict] = []
    for idx, row in enumerate(rows):
        try:
            mapped = _apply_mapping(row, mapping_rules)
            if not mapped.get(join_key):
                failed_details.append({
                    "row_index": idx + 2,  # +2: 表头占第1行，数据从第2行
                    "reason": f"缺失幂等主键 '{join_key}'",
                })
                continue
            mapped_rows.append(mapped)
        except Exception as e:
            failed_details.append({"row_index": idx + 2, "reason": str(e)[:200]})

    if not mapped_rows:
        return {
            "status": "FAILED",
            "total_rows": len(rows),
            "success_count": 0,
            "failed_count": len(rows),
            "failed_details": failed_details,
            "target_table": target_table,
        }

    # upsert 到目标表
    try:
        upsert_result = await upsert_to_target_table(
            db, target_table, mapped_rows, join_key,
        )
    except Exception as e:
        logger.exception("[ucp.excel] upsert failed: %s", e)
        return {
            "status": "FAILED",
            "total_rows": len(rows),
            "success_count": 0,
            "failed_count": len(mapped_rows),
            "failed_details": [{"row_index": -1, "reason": f"目标表写入失败: {str(e)[:300]}"}],
            "target_table": target_table,
        }

    success_count = upsert_result.get("merged_count", 0)
    upsert_failed = upsert_result.get("failed_count", 0)

    # 合并 upsert 失败与映射失败
    if upsert_failed > 0:
        failed_details.append({
            "row_index": -1,
            "reason": f"upsert 阶段失败 {upsert_failed} 行",
        })

    status = "SUCCESS" if (success_count == len(mapped_rows) and not failed_details) else "PARTIAL_SUCCESS"
    if success_count == 0 and upsert_failed > 0:
        status = "FAILED"

    return {
        "status": status,
        "total_rows": len(rows),
        "success_count": success_count,
        "failed_count": len(rows) - success_count,
        "failed_details": failed_details,
        "target_table": target_table,
    }


def _apply_mapping(row: dict, mapping_rules: list[dict] | None) -> dict:
    """应用字段映射：source → target。未在 rules 中的字段保持原名。"""
    if not mapping_rules:
        return dict(row)
    rename_map = {
        r.get("source", ""): r.get("target", "")
        for r in mapping_rules
        if r.get("source") and r.get("target")
    }
    mapped: dict = {}
    for key, value in row.items():
        mapped_key = rename_map.get(key, key)
        mapped[mapped_key] = value
    return mapped


def cleanup_expired_files(max_age_hours: int = 24) -> int:
    """清理超过 max_age_hours 的临时文件，返回清理数量。"""
    import time
    now = time.time()
    count = 0
    for f in UPLOAD_DIR.iterdir():
        if f.is_file() and f.stat().st_mtime < now - max_age_hours * 3600:
            try:
                f.unlink()
                count += 1
            except OSError:
                pass
    return count
