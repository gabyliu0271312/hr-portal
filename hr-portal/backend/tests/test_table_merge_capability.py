"""table_merge.suggest_mapping —— AI 接入 004 底座合规性测试。

覆盖:能力注册元数据、路由存在、输出 deny 扫描(派生表达式不误伤 / SQL 注入命中)、
草稿扫描文本提取。均为纯函数级,不依赖 DB/网络。
"""
import pytest

from app.ai.capabilities import get_capability
from app.ai.policy_guard import AiPolicyError, enforce_output_deny_patterns, validate_capability_policy
from app.main import app
from app.table_tools.engine import eval_derived, _to_field_calls
from app.table_tools.router import _draft_scan_text

CAP_ID = "table_merge.suggest_mapping"


def _route_paths() -> set[str]:
    return {route.path for route in app.routes}


def test_capability_registered_with_expected_metadata():
    cap = get_capability(CAP_ID)
    assert cap is not None
    assert cap.type == "draft"
    assert cap.module == "table_tools"
    assert cap.required_permission == ("table_tools", "E")
    assert "draft_only" in cap.side_effect_tags
    assert cap.confirmation == "none"
    assert cap.model_profile == "fast_json"
    assert cap.sensitive_context == "metadata_only"
    # 只发表头不发明细的策略标记
    assert cap.policy_profile.get("field_context") == "headers_only"
    # sql 已从 deny 移除(避免误伤英文源列名/派生表达式),但代码注入类仍拦
    assert "sql" not in cap.policy_profile.get("deny_patterns", [])
    assert "code" in cap.policy_profile.get("deny_patterns", [])


def test_capability_policy_passes_when_enabled():
    cap = get_capability(CAP_ID)
    decision = validate_capability_policy(cap)
    assert decision.capability_id == CAP_ID
    assert decision.risk_level == "medium"


def test_ai_draft_route_registered():
    assert "/api/v1/table-tools/ai-draft" in _route_paths()


def test_deny_scan_allows_derived_expression():
    """派生表达式 {基数}*{比例} 是合法草稿内容,不得被 deny 误伤。"""
    cap = get_capability(CAP_ID)
    draft = {
        "std_fields": ["养老个人", "公积金个人"],
        "mappings": [
            {
                "column_map": {"基本养老保险(个人)/应缴费额": "养老个人"},
                "derived_fields": [
                    {"target": "公积金个人", "expr": "{缴存基数（元）}*{个人缴存比例}", "round": 2}
                ],
            }
        ],
    }
    # 不抛异常即通过
    assert enforce_output_deny_patterns(cap, _draft_scan_text(draft)) == []


def test_deny_scan_blocks_injection():
    """模型若被诱导输出代码/URL,deny 扫描必须拦截。"""
    cap = get_capability(CAP_ID)
    draft = {
        "std_fields": ["import os; os.system('rm')"],
        "mappings": [],
    }
    with pytest.raises(AiPolicyError):
        enforce_output_deny_patterns(cap, _draft_scan_text(draft))


def test_draft_scan_text_collects_model_output():
    draft = {
        "std_fields": ["养老个人"],
        "mappings": [
            {"column_map": {"源列A": "养老个人"}, "derived_fields": [{"target": "x", "expr": "{a}+{b}"}]}
        ],
    }
    text = _draft_scan_text(draft)
    assert "养老个人" in text
    assert "源列A" in text
    assert "{a}+{b}" in text


# ── 派生字段:复用公共公式引擎(选项2:{列名} 占位 → FIELD) ──────────────

def test_placeholder_converts_to_field_call():
    expr, refs = _to_field_calls("{缴存基数}*{个人缴存比例}")
    assert expr == 'FIELD("缴存基数")*FIELD("个人缴存比例")'
    assert refs == ["缴存基数", "个人缴存比例"]


def test_derived_basic_arithmetic():
    row = {"基数": 10000, "比例": 0.08}
    assert eval_derived("{基数}*{比例}", row.get) == 800.0


def test_derived_uses_engine_functions():
    """彻底改造后派生支持 IF/ROUND/MIN 等全套公式函数(旧 eval_expr 不支持)。"""
    row = {"医疗基数": 80000, "工龄": 12, "基数": 10000}
    assert eval_derived("MIN({医疗基数}*0.08, 5000)", row.get) == 5000.0
    assert eval_derived("IF({工龄}>=10, 1000, 500)", row.get) == 1000
    assert eval_derived("ROUND({基数}*0.08, 2)", row.get) == 800.0


def test_derived_missing_column_returns_none():
    """默认模式保留源行解析的缺失值语义。"""
    row = {"基数": 10000}
    assert eval_derived("{不存在}*2", row.get) is None


def test_derived_missing_fields_are_zero_in_aggregate_mode():
    row = {"B": 100}
    assert eval_derived("{B}+{C}", row.get, missing_as_zero=True) == 100
    assert eval_derived("{B}+{C}", {}.get, missing_as_zero=True) == 0


def test_merge_aggregates_standard_fields_before_derived_formula():
    import io
    import openpyxl
    from app.table_tools.engine import run_merge

    def workbook(headers, values):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "明细"
        ws.append(headers)
        ws.append(values)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    mapping = {
        "name": "来源",
        "sheet_kw": None,
        "header": [1, 1],
        "match": ["员工", "B", "C"],
        "key_map": {"员工": "员工"},
        "column_map": {"B": "B", "C": "C"},
        "derived_fields": [{"target": "A", "expr": "{B}+{C}", "round": 2}],
        "derive_check": None,
        "skip_tokens": [],
    }
    result = run_merge(
        [
            ("B来源.xlsx", workbook(["员工", "B", "C"], ["张三", 100, None])),
            ("C来源.xlsx", workbook(["员工", "B", "C"], ["张三", None, 200])),
        ],
        {"merge_keys": ["员工"], "std_fields": ["B", "C", "A"], "aggregate": "sum"},
        [mapping],
    )
    assert result["rows"] == [{"员工": "张三", "B": 100.0, "C": 200.0, "A": 300.0, "来源": "来源"}]


# ── 源映射维护 ──────────────────────────────────────────────
def test_source_mapping_validation_rejects_short_signature():
    from app.table_tools.models import MergeTemplate
    from app.table_tools.router import SourceMappingIn, _validate_source_mapping

    template = MergeTemplate(name="test", merge_keys=["姓名"], std_fields=["金额"], aggregate="sum")
    payload = SourceMappingIn(
        name="新增映射", match_signature=["姓名", "金额"], source_fields=["姓名", "金额"],
        key_map={"姓名": "姓名"}, column_map={"金额": "金额"},
    )
    with pytest.raises(Exception, match="表头特征至少需要 3 项"):
        _validate_source_mapping(template, payload)


def test_run_merge_reports_unmatched_sheet():
    import io
    import openpyxl
    from app.table_tools.engine import run_merge

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "明细"
    worksheet.append(["姓名", "金额"])
    worksheet.append(["张三", 100])
    buffer = io.BytesIO()
    workbook.save(buffer)
    result = run_merge(
        [("测试.xlsx", buffer.getvalue())],
        {"merge_keys": ["姓名"], "std_fields": ["金额"], "aggregate": "sum"},
        [],
    )
    assert result["anomalies"][0]["type"] == "未匹配源映射"
    assert result["anomalies"][0]["file"] == "测试.xlsx"


def test_run_merge_reports_ambiguous_mapping():
    import io
    import openpyxl
    from app.table_tools.engine import run_merge

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["姓名", "金额", "部门"])
    worksheet.append(["张三", 100, "研发"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    mapping = {
        "sheet_kw": None, "header": [1, 1],
        "match": ["姓名", "金额", "部门"],
        "key_map": {"姓名": "姓名"}, "column_map": {"金额": "金额"},
        "derived_fields": [], "derive_check": None, "skip_tokens": [],
    }
    result = run_merge(
        [("歧义.xlsx", buffer.getvalue())],
        {"merge_keys": ["姓名"], "std_fields": ["金额"], "aggregate": "sum"},
        [{**mapping, "name": "映射A"}, {**mapping, "name": "映射B"}],
    )
    assert result["anomalies"][0]["type"] == "映射命中歧义"


def test_source_mapping_schema_keeps_existing_id():
    from app.table_tools.router import SourceMappingIn

    payload = SourceMappingIn(
        id=42, name="映射", match_signature=["姓名", "证件号", "金额"],
        source_fields=["姓名", "证件号", "金额"],
        key_map={"姓名": "姓名"}, column_map={"金额": "金额"},
    )
    assert payload.id == 42


def test_source_mapping_validation_rejects_blank_name():
    from app.table_tools.models import MergeTemplate
    from app.table_tools.router import SourceMappingIn, _validate_source_mapping

    template = MergeTemplate(name="test", merge_keys=["姓名"], std_fields=["金额"], aggregate="sum")
    payload = SourceMappingIn(
        name="   ", match_signature=["姓名", "证件号", "金额"], source_fields=["姓名", "证件号", "金额"],
        key_map={"姓名": "姓名"}, column_map={"金额": "金额"},
    )
    with pytest.raises(Exception, match="映射名称不能为空"):
        _validate_source_mapping(template, payload)


@pytest.mark.asyncio
async def test_single_mapping_ai_builder_keeps_parent_template_contract(monkeypatch):
    from app.table_tools import ai_builder

    captured = {}

    async def fake_config(_db):
        return "key", None, "model", 30

    async def fake_map(sheet_info, std_fields, merge_keys, *_args):
        captured.update(sheet_info=sheet_info, std_fields=std_fields, merge_keys=merge_keys)
        return {"key_map": {"员工姓名": "姓名"}, "column_map": {"个人缴存": "公积金个人"},
                "derived_fields": [], "_confidence": 0.9, "_notes": ""}

    monkeypatch.setattr(ai_builder, "_get_ai_config", fake_config)
    monkeypatch.setattr(ai_builder, "_step2_map_sheet", fake_map)
    result = await ai_builder.build_mapping_draft(
        {"file": "样表.xlsx", "sheet": "明细", "columns": ["员工姓名", "个人缴存"]},
        ["公积金个人"], ["姓名"], "", object(),
    )
    assert result["column_map"] == {"个人缴存": "公积金个人"}
    assert captured["std_fields"] == ["公积金个人"]
    assert captured["merge_keys"] == ["姓名"]
    assert captured["sheet_info"]["columns"] == ["员工姓名", "个人缴存"]


def test_merge_runs_derive_check_after_aggregation():
    from app.table_tools.engine import aggregate_records, _apply_derive_checks

    rows, anomalies = aggregate_records(
        [
            ({"员工": "张三", "B": 100}, "B来源"),
            ({"员工": "张三", "C": 200, "总额": 300}, "C来源"),
        ],
        ["员工"], ["B", "C", "总额"],
    )
    _apply_derive_checks(
        rows, ["员工"],
        [{"sum_of": ["B", "C"], "equals_col": "总额", "tol": 0.05}],
        anomalies,
    )
    assert anomalies == []


def test_merge_rejects_inconsistent_derived_formula():
    from app.table_tools.engine import _merge_derived_fields

    with pytest.raises(ValueError, match="派生字段 A 在多个来源映射中的公式不一致"):
        _merge_derived_fields([
            {"derived_fields": [{"target": "A", "expr": "{B}+{C}", "round": 2}]},
            {"derived_fields": [{"target": "A", "expr": "{B}-{C}", "round": 2}]},
        ])


def test_derived_field_does_not_override_direct_mapping_from_other_source():
    """表1 派生 c=a*b,表2 直接映射字段3→c,两人不同:表2 的 c 应保留直接映射值。"""
    import io
    import openpyxl
    from app.table_tools.engine import run_merge

    def workbook(headers, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "明细"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    mapping1 = {
        "name": "表1",
        "sheet_kw": None,
        "header": [1, 1],
        "match": ["员工", "字段1", "字段2"],
        "key_map": {"员工": "员工"},
        "column_map": {"字段1": "a", "字段2": "b"},
        "derived_fields": [{"target": "c", "expr": "{a}*{b}", "round": 2}],
        "derive_check": None,
        "skip_tokens": [],
    }
    mapping2 = {
        "name": "表2",
        "sheet_kw": None,
        "header": [1, 1],
        "match": ["员工", "字段3", "备注"],
        "key_map": {"员工": "员工"},
        "column_map": {"字段3": "c"},
        "derived_fields": [],
        "derive_check": None,
        "skip_tokens": [],
    }
    result = run_merge(
        [
            ("表1.xlsx", workbook(["员工", "字段1", "字段2"], [["张三", 10, 20]])),
            ("表2.xlsx", workbook(["员工", "字段3", "备注"], [["李四", 999, "无"]])),
        ],
        {"merge_keys": ["员工"], "std_fields": ["a", "b", "c"], "aggregate": "sum"},
        [mapping1, mapping2],
    )
    assert result["rows"] == [
        {"员工": "张三", "a": 10.0, "b": 20.0, "c": 200.0, "来源": "表1"},
        {"员工": "李四", "a": 0, "b": 0, "c": 999.0, "来源": "表2"},
    ]


def test_derived_field_overrides_direct_mapping_when_inputs_present():
    """同人同时出现在表1(派生 a*b)与表2(直接映射 c):派生优先,覆盖直接映射值。"""
    import io
    import openpyxl
    from app.table_tools.engine import run_merge

    def workbook(headers, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "明细"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    mapping1 = {
        "name": "表1",
        "sheet_kw": None,
        "header": [1, 1],
        "match": ["员工", "字段1", "字段2"],
        "key_map": {"员工": "员工"},
        "column_map": {"字段1": "a", "字段2": "b"},
        "derived_fields": [{"target": "c", "expr": "{a}*{b}", "round": 2}],
        "derive_check": None,
        "skip_tokens": [],
    }
    mapping2 = {
        "name": "表2",
        "sheet_kw": None,
        "header": [1, 1],
        "match": ["员工", "字段3", "备注"],
        "key_map": {"员工": "员工"},
        "column_map": {"字段3": "c"},
        "derived_fields": [],
        "derive_check": None,
        "skip_tokens": [],
    }
    result = run_merge(
        [
            ("表1.xlsx", workbook(["员工", "字段1", "字段2"], [["张三", 10, 20]])),
            ("表2.xlsx", workbook(["员工", "字段3", "备注"], [["张三", 999, "无"]])),
        ],
        {"merge_keys": ["员工"], "std_fields": ["a", "b", "c"], "aggregate": "sum"},
        [mapping1, mapping2],
    )
    assert result["rows"] == [
        {"员工": "张三", "a": 10.0, "b": 20.0, "c": 200.0, "来源": "表1 + 表2"},
    ]
